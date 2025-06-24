from tkinter import *
from tkinter import messagebox
from PIL import Image, ImageTk
from PIL import Image as PILImage
import customtkinter
import customtkinter as ctk
from tkinter import ttk
import tkinter as tk
from tkcalendar import DateEntry
from datetime import datetime, timedelta
import locale
from Database.Conexion_SQL_Server import *
from Database.Datos import * 
import sys
from Generacion_Reportes import generar_reporte_pdf
from openpyxl import load_workbook
from tkinter import filedialog
from copy import copy
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import matplotlib.pyplot as plt
from io import BytesIO
from reportlab.platypus import Image

ventana = customtkinter.CTk()
ventana.title("Factor Ajuste")
ventana.config(bg="#C0C0C0")
customtkinter.set_appearance_mode("dark")

# Obtener el tamaño de la pantalla
screen_width = ventana.winfo_screenwidth()
screen_height = ventana.winfo_screenheight()

# Configurar la ventana para que ocupe todo el tamaño de la pantalla
ventana.geometry(f"{screen_width}x{screen_height}")

# Escalar fuente
fuente_base = int(screen_width * 0.015)

# Poner los elementos del programa en Español
locale.setlocale(locale.LC_TIME, "es_MX")

# sys.argv contiene los argumentos pasados
# RPE = sys.argv[1]  # Primer argumento
# RPU = sys.argv[2]    # Segundo argumento
RPU = 272680903204
RPE = "JA117"

# 🔹 Cargar imagen correctamente con PIL
imagen = PILImage.open("Imagenes/Logo_CFE.png").resize((int(screen_width*0.3), int(screen_height*0.1))).convert("RGBA")
imagen_tk = ImageTk.PhotoImage(imagen)

imagen2 = PILImage.open("Imagenes/Logo_Medicion.jpg").resize((int(screen_width*0.15), int(screen_height*0.16))).convert("RGBA")
imagen_tk2 = ImageTk.PhotoImage(imagen2)

# 🔹 Encabezado
encabezado = customtkinter.CTkFrame(ventana, fg_color="White", bg_color="Black")
encabezado.place(relx=0, rely=0, relwidth=1, relheight=0.2)

img = customtkinter.CTkLabel(encabezado, image=imagen_tk, text=" ")
img.place(relx=0.025, rely=0.2)
img.image = imagen_tk

titulo = customtkinter.CTkLabel(encabezado, text="COMISION FEDERAL DE ELECTRICIDAD \n DIVISION CENTRO ORIENTE \n ZONA DE DISTRIBUCION TULA",
                                    bg_color="White", fg_color="White", text_color="Black", font=("Arial", int(fuente_base * 0.9), "bold"))
titulo.place(relx=0.39, rely=0.2)

img2 = customtkinter.CTkLabel(encabezado, image=imagen_tk2, text=" ")
img2.place(relx=0.825, rely=0.08)
img2.image = imagen_tk2
# Fin de Encabezado

# Datos Generales
# Datos RPE
ListaRPE = {
"9L99B": "Francisco Javier Velazquez Cervantes",
"JA117": "Jerzael Abisai Mera Zapatero"
}

datos = customtkinter.CTkFrame(ventana, fg_color="green", bg_color="Black")
datos.place(relx=0, rely=0.15, relwidth=1, relheight=0.12)

estiloDatos = { 
        "fg_color": ("Green"), 
        "bg_color": ("Green"), 
        "text_color": ("White"), 
        "width": 50, 
        "height": 20, 
        "font": ("Arial", int(fuente_base * 0.65),"bold")
}

def ExtraccionCuenta():
    texto = cuenta.cget("text")  # Obtener el texto del Label

    Cclo = texto[:2]        # Primeras 2 letras
    pblacion = texto[7:9]   # Caracteres 6 y 7
    rta = texto[9:12]       # 3 caracteres después
    flio = texto[12:16]       # Últimos 4 caracteres

    ciclo.configure(text=Cclo)
    poblacion.configure(text=pblacion)
    ruta.configure(text=rta)
    folio.configure(text=flio)

ahora = datetime.now()
fecha_hora = ahora.strftime("%A, %d %B, %Y")

lbl1 = customtkinter.CTkLabel(datos, text="R.P.E :", **estiloDatos)
lbl1.place(relx=0.05, rely=0.2)

rpe = customtkinter.CTkLabel(datos, text=RPE, **estiloDatos)
rpe.place(relx=0.1, rely=0.2)

lbl2 = customtkinter.CTkLabel(datos, text="Nombre :", **estiloDatos)
lbl2.place(relx=0.25, rely=0.2)

Nomrpe = customtkinter.CTkLabel(datos, text="Nombre del RPE", **estiloDatos)
Nomrpe.place(relx=0.3, rely=0.2)

FechaAct = customtkinter.CTkLabel(datos, text=fecha_hora, **estiloDatos)
FechaAct.place(relx=0.84, rely=0.2)

# Buscar el RPE dentro del diccionario de trabajadores
nombre_rpe = ListaRPE.get(RPE, "Trabajador no encontrado")  # Si no está en el diccionario, muestra "Trabajador no encontrado"
    
Nomrpe.configure(text=nombre_rpe)

# Datos RPU - Fila 1
DatosRPU = BD.datosRPU(RPU)
MedidoresRPU = BD.medidores(RPU)

lbl3 = customtkinter.CTkLabel(datos, text="R.P.U :", **estiloDatos)
lbl3.place(relx=0.05, rely=0.4)

rpu = customtkinter.CTkLabel(datos, text=RPU, **estiloDatos)
rpu.place(relx=0.1, rely=0.4)

lbl4 = customtkinter.CTkLabel(datos, text="Cuenta :", **estiloDatos)
lbl4.place(relx=0.25, rely=0.4)

cuenta = customtkinter.CTkLabel(datos, text="----", **estiloDatos)
cuenta.place(relx=0.3, rely=0.4)

lbl5 = customtkinter.CTkLabel(datos, text="Nombre :", **estiloDatos)
lbl5.place(relx=0.45, rely=0.4)

Nomrpu = customtkinter.CTkLabel(datos, text="Nombre del RPU", **estiloDatos)
Nomrpu.place(relx=0.5, rely=0.4)

lbl6 = customtkinter.CTkLabel(datos, text="Medidor :", **estiloDatos)
lbl6.place(relx=0.65, rely=0.4)

medidores = [registro[1] for registro in MedidoresRPU]
selected_medidor = ctk.StringVar()
selected_medidor.set(medidores[0] if medidores else "")

medidor = customtkinter.CTkOptionMenu(datos, variable=selected_medidor ,values=medidores, **estiloDatos)
medidor.place(relx=0.7, rely=0.4)

# Datos RPU - Fila 2
direccion = ""

lbl7 = customtkinter.CTkLabel(datos, text="Ciclo :",  **estiloDatos)
lbl7.place(relx=0.05, rely=0.6)

ciclo = customtkinter.CTkLabel(datos, text=" ", **estiloDatos)
ciclo.place(relx=0.1, rely=0.6)

lbl8 = customtkinter.CTkLabel(datos, text="Poblacion :", **estiloDatos)
lbl8.place(relx=0.25, rely=0.6)

poblacion = customtkinter.CTkLabel(datos, text=" ", **estiloDatos)
poblacion.place(relx=0.3, rely=0.6)

lbl9 = customtkinter.CTkLabel(datos, text="Ruta :", **estiloDatos)
lbl9.place(relx=0.45, rely=0.6)

ruta = customtkinter.CTkLabel(datos, text=" ", **estiloDatos)
ruta.place(relx=0.5, rely=0.6)

lbl10 = customtkinter.CTkLabel(datos, text="Folio :", **estiloDatos)
lbl10.place(relx=0.65, rely=0.6)

folio = customtkinter.CTkLabel(datos, text=" ", **estiloDatos)
folio.place(relx=0.7, rely=0.6)

lbl11 = customtkinter.CTkLabel(datos, text="Agencia :", **estiloDatos)
lbl11.place(relx=0.85, rely=0.6)

agencia = customtkinter.CTkLabel(datos, text="Nombre de la Agencia", **estiloDatos)
agencia.place(relx=0.9, rely=0.6)

# Diccionario de las agencias
AGENCIAS = {
    "A": "Progreso",
    "B": "Tula",
    "C": "Tlaxcoapan",
    "D": "Tepeji"
}

# Cargar Datos de RPU de la BD en las etiquetas

if DatosRPU and len(DatosRPU) >= 4:
    Nomrpu.configure(text=DatosRPU[1])  # Nombre RPU
    direccion = DatosRPU[2]  # Nombre RPU
    cuenta.configure(text=DatosRPU[3])  # Cuenta

     # Obtener la agencia del resultado
    codigo_agencia = DatosRPU[4]
    
    # Convertir el código a su nombre correspondiente
    nombre_agencia = AGENCIAS.get(codigo_agencia, "Desconocida")  # Si no está en el diccionario, muestra "Desconocida"
    
    agencia.configure(text=nombre_agencia)
else:
    # Si no hay datos, mostrar un mensaje de "No encontrado"
    Nomrpu.configure(text="No encontrado")
    direccion = None
    cuenta.configure(text="No encontrado")
    agencia.configure(text="No encontrado")

# Fin de Datos Generales

# Llamar a la función inmediatamente al inicio
ExtraccionCuenta()

# Seccion de Calculos
calculos = customtkinter.CTkFrame(ventana, fg_color="gray", bg_color="Black")
calculos.place(relx=0, rely=0.27, relwidth=0.35, relheight=0.75)

lbl12 = customtkinter.CTkLabel(calculos, text="BASE DE CALCULO PARA REALIZAR AJUSTE", fg_color="Gray", bg_color="Gray", text_color="White", font=("Arial", int(fuente_base * 0.7),"bold"))
lbl12.place(relx=0.2, rely=0.02)

# Tabla
class TablaApp(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color="#2b2b2b", bg_color="#2b2b2b")
        self.place(relx=0.02, rely=0.07, relwidth=0.96, relheight=0.09)

        # Canvas ocupa todo el Frame
        self.canvas = ctk.CTkCanvas(self, bg="#2b2b2b", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)

        # Frame que contiene la tabla
        self.table_frame = ctk.CTkFrame(self.canvas, fg_color="#2b2b2b", bg_color="#2b2b2b")
        self.table_window = self.canvas.create_window((0, 0), window=self.table_frame, anchor="nw")

        # Configurar columnas para que se repartan equitativamente
        for i in range(5):
            self.table_frame.columnconfigure(i, weight=1)

        self.table = []
        self.create_table()

        # Vincula el redimensionamiento
        self.bind("<Configure>", self.resize_table_frame)
        self.table_frame.bind("<Configure>", self.on_frame_configure)

    def create_table(self):
        headers = ["Corriente", "Voltaje", "kVA real", "kVA cronometro", "% Registración"]
        for col, header in enumerate(headers):
            label = ctk.CTkLabel(self.table_frame,
                                 text=header,
                                 fg_color="green",
                                 bg_color="green",
                                 text_color="white",
                                 font=("Arial", int(fuente_base * 0.55), "bold"))
            label.grid(row=0, column=col, sticky="nsew", padx=1, pady=1)

        self.add_row()

    def add_row(self):
        row = []
        for col in range(5):
            entry = ctk.CTkEntry(self.table_frame,
                                 fg_color="white",
                                 bg_color="white",
                                 border_color="white",
                                 text_color="black")
            entry.grid(row=1, column=col, sticky="nsew", padx=1, pady=1)
            row.append(entry)
        self.table.append(row)

    def resize_table_frame(self, event):
        # Hacer que el frame de la tabla siempre tenga el mismo tamaño que el canvas
        self.canvas.itemconfig(self.table_window, width=event.width)

    def on_frame_configure(self, event):
        # Asegura que el canvas ajuste la región visible correctamente
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
    
tabla = TablaApp(calculos)
# Fin de la tabla

# Apartado del Periodo
lbl13 = customtkinter.CTkLabel(calculos, text="Periodo", fg_color="Gray", bg_color="Gray", text_color="White", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl13.place(relx=0.05, rely=0.18)

class PeriodoSelector(ctk.CTkFrame):
    def __init__(self, master=None):
        super().__init__(master)
        self.place(relwidth=0.9, relheight=0.15)
        self.pack_propagate(False)

        # Label Desde
        self.label_desde = ctk.CTkLabel(self, text="Desde (Año-Mes-Día):")
        self.label_desde.place(relx=0.2, rely=0.1)

        # Selector de fecha "Desde" (Formato completo)
        self.date_desde = DateEntry(self, background='darkgreen', 
                                    foreground='white', borderwidth=2, year=2025, 
                                    date_pattern="yyyy-mm-dd", font=(int(fuente_base * 0.35)))  
        self.date_desde.place(relx=0.23, rely=0.5, relwidth=0.22, relheight=0.25)

        # Label Hasta
        self.label_hasta = ctk.CTkLabel(self, text="Hasta (Año-Mes-Día):")
        self.label_hasta.place(relx=0.55, rely=0.1)

        # Selector de fecha "Hasta" (Formato completo)
        self.date_hasta = DateEntry(self, background='darkgreen', 
                                    foreground='white', borderwidth=2, year=2025, 
                                    date_pattern="yyyy-mm-dd", font=(int(fuente_base * 0.35)))  
        self.date_hasta.place(relx=0.57, rely=0.5, relwidth=0.22, relheight=0.25)

    def get_range(self):
        # Obtener la fecha seleccionada y mostrarla con Año-Mes-Día
        desde_año = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%Y")
        desde_mes = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%m")
        desde_mes_palabra = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%B")
        desde_dia = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%d")

        hasta_año = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%Y")
        hasta_mes = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%m")
        hasta_mes_palabra = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%B")
        hasta_dia = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%d")
        
        return desde_año, desde_mes, desde_mes_palabra, desde_dia, hasta_año, hasta_mes, hasta_mes_palabra, hasta_dia


periodo_selector = PeriodoSelector(calculos)
periodo_selector.place(relx=0.05, rely=0.22)  # Ajusta la posición
# Fin del Apartado Periodo

lbl14 = customtkinter.CTkLabel(calculos, text="Anomalia :", fg_color="Gray", bg_color="Gray", text_color="White", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl14.place(relx=0.05, rely=0.40)

anomalia = customtkinter.CTkEntry(calculos, fg_color="White", bg_color="Gray", text_color="Black")
anomalia.place(relx=0.2, rely=0.40, relwidth=0.22, relheight=0.05)

class PeriodoIncompleto(ctk.CTkFrame):
    def __init__(self, master=None, periodo_selector=None):
        super().__init__(master)
        self.place(relwidth=0.9, relheight=0.15)
        self.pack_propagate(False)

        self.periodo_selector = periodo_selector  # Guardamos la referencia del PeriodoSelector

        # Label Desde
        self.label_desde = ctk.CTkLabel(self, text="Desde (Año-Mes-Día):")
        self.label_desde.place(relx=0.22, rely=0.1)

        # Selector de fecha "Desde" (Formato completo)
        self.date_desde = DateEntry(self, background='darkred', 
                                    foreground='white', borderwidth=2, year=2025, 
                                    date_pattern="yyyy-mm-dd", font=(int(fuente_base * 0.35)))  
        self.date_desde.place(relx=0.23, rely=0.5, relwidth=0.22, relheight=0.25)

        # Label Hasta
        self.label_hasta = ctk.CTkLabel(self, text="Hasta (Año-Mes-Día):")
        self.label_hasta.place(relx=0.55, rely=0.1)

        # Selector de fecha "Hasta" (Formato completo)
        self.date_hasta = DateEntry(self, background='darkred', 
                                    foreground='white', borderwidth=2, year=2025, 
                                    date_pattern="yyyy-mm-dd", font=(int(fuente_base * 0.35)))  
        self.date_hasta.place(relx=0.57, rely=0.5, relwidth=0.22, relheight=0.25)

    def get_range(self):
       
        # Obtener la fecha seleccionada y mostrarla con Año-Mes-Día
        incompleto_desde_año = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%Y")
        incompleto_desde_mes = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%m")
        incompleto_desde_mes_palabra = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%B")
        incompleto_desde_dia = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%d")

        incompleto_hasta_año = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%Y")
        incompleto_hasta_año_formato = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%Y")[-2:]
        incompleto_hasta_mes = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%m")
        incompleto_hasta_mes_palabra = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%B")
        incompleto_hasta_dia = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%d")
        
        return incompleto_desde_año, incompleto_desde_mes, incompleto_desde_mes_palabra, incompleto_desde_dia, incompleto_hasta_año, incompleto_hasta_año_formato, incompleto_hasta_mes, incompleto_hasta_mes_palabra, incompleto_hasta_dia


lbl34 = customtkinter.CTkLabel(calculos, text="Periodo Incompleto", fg_color="Gray", bg_color="Gray", text_color="White", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl34.place(relx=0.05, rely=0.48)

periodo_selector_incompleto = PeriodoIncompleto(calculos, periodo_selector)
periodo_selector_incompleto.place(relx=0.05, rely=0.53)  # Ajusta la posición
# Fin del apartado de Periodo Incompleto

# Apartado para Lectura Inicial y Lectura Final
lbl35 = customtkinter.CTkLabel(calculos, text="Lectura Inicial", fg_color="Gray", bg_color="Gray", text_color="White", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl35.place(relx=0.2, rely=0.71)

LectInic = customtkinter.CTkEntry(calculos, fg_color="White", bg_color="Gray", text_color="Black")
LectInic.place(relx=0.19, rely=0.76, relwidth=0.2, relheight=0.05)

lbl36 = customtkinter.CTkLabel(calculos, text="Lectura Final", fg_color="Gray", bg_color="Gray", text_color="White", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl36.place(relx=0.6, rely=0.71)

LectFin = customtkinter.CTkEntry(calculos, fg_color="White", bg_color="Gray", text_color="Black")
LectFin.place(relx=0.59, rely=0.76, relwidth=0.2, relheight=0.05)
# Fin del apartado para Lectura Inicial y Lectura Final

def Calculos():
    Tabla.bind("<Button-1>", toggle_checkbox)

    columnas = [0] * 5
    conteo = [0] * 5  # Para contar cuántos valores válidos hay en cada columna

    for row in tabla.table:
        for col, entry in enumerate(row):
            try:
                value = entry.get().strip()
                if value:
                    columnas[col] += float(value)
                    conteo[col] += 1
            except ValueError:
                pass # Ignorar valores no numéricos

    # Calcular promedios con manejo de división por cero
    factordeajuste = (1 / columnas[4]) * 100 if columnas[4] > 0 else 0

    # Asignar valores a las etiquetas correspondientes
    corriente.configure(text=f"{columnas[0]:.1f}")
    voltaje.configure(text=f"{columnas[1]:.1f}")
    kvareal.configure(text=f"{columnas[2]:.4f}")
    kvacronometro.configure(text=f"{columnas[3]:.4f}")
    registracion.configure(text=f"{columnas[4]:.3f}")
    FactAjus.configure(text=f"{factordeajuste:.4f}")

    # Limpiar la tabla antes de insertar nuevos datos
    Tabla.delete(*Tabla.get_children())

    # Variables para obtener los cálculos
    suma_dias = 0
    suma_kWh_total = 0
    suma_kWh_total_DF = 0
    suma_kWh_total_D = 0

    MedSelect = selected_medidor.get()

    tarifa = BD.tarifa(RPU, MedSelect)
    lbl25.configure(text=tarifa[0])

    # Mostrar los datos en la tabla

    desde_año, desde_mes, desde_mes_palabra, desde_dia, hasta_año, hasta_mes, hasta_mes_palabra, hasta_dia = periodo_selector.get_range()
    desde = f"{desde_año}-{desde_mes}-{desde_dia}"
    hasta = f"{hasta_año}-{hasta_mes}-{hasta_dia}"
    
    for row in BD.mostrardatos(RPU, desde, hasta):
        FECHA, DESDE, HASTA, CONSUMO = row

        # Convertir CONSUMO a entero si es necesario
        if isinstance(CONSUMO, str):
            CONSUMO = int(CONSUMO)

        # Formatear la fecha para mostrar los últimos 2 dígitos del año y los 2 dígitos del mes
        FECHA_FORMATO = FECHA.strftime('%y%m')

        # Formatear las fechas desde y hasta para mostrar solo la fecha y no la hora
        DESDE_FORMATO = DESDE.strftime('%Y-%m-%d')
        HASTA_FORMATO = HASTA.strftime('%Y-%m-%d')

        # Calcular nuevas columnas
        DIAS = (HASTA - DESDE).days
        
        # Insertar en la tabla en orden cronológico
        Tabla.insert("", "end", values=(FECHA_FORMATO, DESDE_FORMATO, HASTA_FORMATO, DIAS, CONSUMO, " ", " ", "✓"))
    
    totalDias.configure(text=suma_dias)
    totalCFRkwh.configure(text=suma_kWh_total)
    totalDFkwh.configure(text=suma_kWh_total_DF)
    totalDkwh.configure(text=suma_kWh_total_D)
    lbl28.configure(text=suma_kWh_total_D)

    # Diccionario de anomalias
    Anomalias = {
        "EF01": "DIRECTO CON ESTIMACION ERRÓNEA",
        "EF02": "TARIFA INCORRECTA",
        "EF03": "SERVICIO CON LECTURAS ESTIMADAS",
        "EF04": "SERVICIO CON MEDICIÓN Y NO SE FACTURA",
        "EF05": "SERVICIO DE CONCURRENCIA DE TARIFAS",
        "EF06": "LECTURA MAL TOMADA",
        "EF07": "CARGA MAYOR A LA CONTRATADA",
        "EF08": "MULTIPLICADOR DE LECTURAS ERRÓNEO",
        "EF09": "CLAVE DE TIPO DE SUMINISTRO ERRÓNEA",

        "FM01": "SERVICIO DIRECTO",
        "FM02": "MULTIPLICADOR DE LECTURAS ERRÓNEO",
        "FM03": "MEDIDOR MAL CONECTADO",
        "FM04": "BOBINA DE POTENCIAL ABIERTA",
        "FM05": "MEDIDOR QUEMADO",
        "FM06": "TRANSFORMADOR DE INSTRUMENTO DAÑADO",
        "FM07": "TRANSFORMADOR DE INSTRUMENTO MAL CONTECTADO",
        "FM08": "CIRCUTIO DE CORRIENTE DE T.C A TIERRA",
        "FM09": "REGISTRO SIN ENGRANAR",
        "FM10": "NÚMERO DE MEDIDOR DIFERENTE AL DE FACTURACION",
        "FM11": "CLAVE DE TIPO DE SUMINISTRO MAL ERRÓNEA",
        "FM12": "MEDICION INADECUADA AL CIRCUITO",
        "FM13": "MEDIDOR DESTRUIDO",
        "FM14": "ALAMBRADO SECUNDARIO DAÑADO",
        "FM15": "SECUDNARIO DE T.C. EN CORTO CIRCUITO",
        "FM16": "MEDIDOR MAL PROGRAMADO",
        "FM17": "REGISTRO ATORADO",
        "FM18": "DISCO ATORADO",

        "UI01": "SELLOS VIOLADOS",
        "UI02": "SELLO FALSIFICADO",
        "UI03": "SERVICIO DIRECTO SIN CONTRATO",
        "UI04": "CARGA CONECTADA ANTES DE LA MEDICIÓN",
        "UI05": "MANECILLAS INVERTIDAS",
        "UI06": "DISCO ATORADO",
        "UI07": "AJUSTE DE MEDIDORES MOVIDOS",
        "UI08": "RETORNO DE CORRIENTE ABIERTO",
        "UI09": "INTERVENCIÓN DE LAS CONEXIONES DEL EQUIPO",
        "UI10": "ALDABAS DE POTENCILA ABIERTAS",
        "UI11": "MEDIDOR INVERTIDO",
        "UI12": "TERMINALES DE BASE ENCHUFES PUENTEADAS",
        "UI13": "SIN CONTRATO CON MEDICIÓN INSTALADA",
        "UI14": "PASA ENERGÍA A OTRO DOMICILIO",
        "UI15": "ALTERACIÓN A LA PROGRAMACIÓN DEL MEDIDOR",
        "UI16": "RETIRA SU MEDIDOR Y COLOCA OTRO QUE NO FACTURA",
        "UI17": "REGISTRO DEL MEDIDOR ALTERADO O CAMBIADO",
        "UI18": "OTROS USOS INDEBIDOS"
    }

    codigo_anomalia = anomalia.get()
    nombre_anomalia = Anomalias.get(codigo_anomalia, "Desconocida")  # Si no está en el diccionario, muestra "Desconocida"
    lbl33.configure(text=nombre_anomalia)

    ultimo_periodo_incompleto["desde"] = 0
    ultimo_periodo_incompleto["hasta"] = 0
    ultimo_resumen_incompleto["dias"] = 0
    ultimo_resumen_incompleto["df"] = 0
    ultimo_resumen_incompleto["d"] = 0

    agregar.configure(state="disabled")
    btnConfirmar.configure(state="normal")

    return suma_dias, suma_kWh_total, suma_kWh_total_DF, suma_kWh_total_D, rpu, desde, hasta

cargar = ctk.CTkButton(calculos, text="Calcular",fg_color="#2b2b2b", bg_color="Gray", font=("Arial", int(fuente_base * 0.6), "bold"), hover_color="Green", command=Calculos)
cargar.place(relx=0.6, rely=0.40, relwidth=0.2, relheight=0.06)

def confirmar_calculo():
     # Variables acumuladoras
    suma_dias = 0
    suma_kWh_total = 0
    suma_kWh_total_DF = 0
    suma_kWh_total_D = 0

    Tabla.unbind("<Button-1>")  # Desactiva el click sobre la tabla para evitar más selecciones

    try:
        FA = float(FactAjus.cget("text"))
    except ValueError:
        FA = 0    

    for item in Tabla.get_children():
        valores = list(Tabla.item(item, "values"))

        if valores[7] == "✓":
            try:
                FECHA_FORMATO = valores[0]
                DESDE = datetime.strptime(valores[1], '%Y-%m-%d')
                HASTA = datetime.strptime(valores[2], '%Y-%m-%d')
                DIAS = int(valores[3])
                CONSUMO = int(valores[4])  # Ya viene en la tabla

                CONSUMO_DF = round(FA * CONSUMO)  # Redondear hacia arriba o abajo
                CONSUMO_D = round(CONSUMO_DF - CONSUMO)  # Redondear hacia arriba o abajo

                # Reemplazar valores, manteniendo los anteriores y agregando los nuevos
                Tabla.item(item, values=(FECHA_FORMATO, DESDE.strftime('%Y-%m-%d'), HASTA.strftime('%Y-%m-%d'), DIAS, CONSUMO, CONSUMO_DF, CONSUMO_D, "✓" ))

                # Acumular en variables
                suma_dias += DIAS
                suma_kWh_total += CONSUMO
                suma_kWh_total_DF += CONSUMO_DF
                suma_kWh_total_D += CONSUMO_D

            except Exception as e:
                print(f"Error al calcular fila: {e}")
        else:
            Tabla.delete(item)  # Eliminar filas no seleccionadas

    # Actualizar etiquetas con resultados
    totalDias.configure(text=suma_dias)
    totalCFRkwh.configure(text=suma_kWh_total)
    totalDFkwh.configure(text=suma_kWh_total_DF)
    totalDkwh.configure(text=suma_kWh_total_D)
    lbl28.configure(text=suma_kWh_total_D)

    # Obtener fechas del periodo_selector
    items = Tabla.get_children()
    if not items:
        messagebox.showerror("Error", "⚠ No hay datos en la tabla para validar.")
        return False
    
    # Tomar el primer y último registro de la tabla
    primer_registro = Tabla.item(items[0], "values")  # Primer registro
    ultimo_registro = Tabla.item(items[-1], "values")  # Último registro

    desde_año_txt = datetime.strptime(primer_registro[1], "%Y-%m-%d").strftime("%Y")
    desde_mes_palabra_txt = datetime.strptime(primer_registro[1], "%Y-%m-%d").strftime("%B")
    desde_dia_txt = datetime.strptime(primer_registro[1], "%Y-%m-%d").strftime("%d")

    hasta_año_txt = datetime.strptime(ultimo_registro[2], "%Y-%m-%d").strftime("%Y")
    hasta_mes_palabra_txt = datetime.strptime(ultimo_registro[2], "%Y-%m-%d").strftime("%B")
    hasta_dia_txt = datetime.strptime(ultimo_registro[2], "%Y-%m-%d").strftime("%d")

    lbl31.configure(text=f"Del {desde_dia_txt} de {desde_mes_palabra_txt} de {desde_año_txt} al {hasta_dia_txt} de {hasta_mes_palabra_txt} de {hasta_año_txt}")

    btnConfirmar.configure(state="disabled")
    agregar.configure(state="normal")

# Guardar último período incompleto agregado
ultimo_periodo_incompleto = {"desde": None, "hasta": None}
ultimo_resumen_incompleto = {"dias": 0, "df": 0, "d": 0}

def Agregar():
    global ultimo_periodo_incompleto, ultimo_resumen_incompleto

    # Obtener fechas del periodo incompleto
    incompleto_desde_año, incompleto_desde_mes, _, incompleto_desde_dia, \
    incompleto_hasta_año, incompleto_hasta_año_formato, incompleto_hasta_mes, _, incompleto_hasta_dia = \
        periodo_selector_incompleto.get_range()

    desde_incompleto = f"{incompleto_desde_año}-{incompleto_desde_mes}-{incompleto_desde_dia}"
    hasta_incompleto = f"{incompleto_hasta_año}-{incompleto_hasta_mes}-{incompleto_hasta_dia}"
    fecha_incompleto = f"{incompleto_hasta_año_formato}{incompleto_hasta_mes}"

    # Si es el mismo periodo que el anterior, no hacer nada
    if (ultimo_periodo_incompleto["desde"] == desde_incompleto and
        ultimo_periodo_incompleto["hasta"] == hasta_incompleto):
        messagebox.showinfo("Sin cambios", "Mismo periodo incompleto agregado, no se ha modificado los datos de la tabla.")
        return

    # ✅ Restar el periodo anterior de las sumas totales
    if ultimo_resumen_incompleto["dias"] > 0:
        suma_actual_dias = int(totalDias.cget("text")) - ultimo_resumen_incompleto["dias"]
        suma_actual_df = int(totalDFkwh.cget("text")) - ultimo_resumen_incompleto["df"]
        suma_actual_d = int(totalDkwh.cget("text")) - ultimo_resumen_incompleto["d"]

        totalDias.configure(text=suma_actual_dias)
        totalDFkwh.configure(text=suma_actual_df)
        totalDkwh.configure(text=suma_actual_d)
        lbl28.configure(text=suma_actual_d)

    # Eliminar el periodo incompleto anterior en la tabla
    for item_id in Tabla.get_children():
        valores = Tabla.item(item_id)["values"]
        if valores and valores[7] == " ":
            Tabla.delete(item_id)
            break

    # Validar fechas
    if not validar_fecha(Tabla, desde_incompleto, hasta_incompleto):
        return

    # Convertir a fecha y calcular días
    desde_dt = datetime.strptime(desde_incompleto, "%Y-%m-%d").date()
    hasta_dt = datetime.strptime(hasta_incompleto, "%Y-%m-%d").date()
    dias_incompletos = (hasta_dt - desde_dt).days

    # Obtener datos de cálculo
    try:
        FA = float(FactAjus.cget("text"))
        LI = int(LectInic.get())
        LF = int(LectFin.get())
    except ValueError:
        FA = 0
        LI = 0
        LF = 0

    diferencia = abs(LF - LI)
    consumo_incompleto_DF = round(diferencia * FA)
    consumoDF_incompleto = consumo_incompleto_DF

    # Insertar nueva fila
    insertar_en_orden(Tabla, fecha_incompleto,
                      (fecha_incompleto, desde_incompleto, hasta_incompleto, dias_incompletos, 0,
                       consumo_incompleto_DF, consumoDF_incompleto, " "))

    # Actualizar fechas en el label inferior
    fechas = []
    for item_id in Tabla.get_children():
        valores = Tabla.item(item_id)["values"]
        if len(valores) >= 3:
            f_inicio = datetime.strptime(valores[1], "%Y-%m-%d").date()
            f_fin = datetime.strptime(valores[2], "%Y-%m-%d").date()
            fechas.extend([f_inicio, f_fin])

    if fechas:
        f_ini = min(fechas)
        f_fin = max(fechas)
        lbl31.configure(text=f"Del {f_ini.strftime('%d')} de {f_ini.strftime('%B')} de {f_ini.year} "
                             f"al {f_fin.strftime('%d')} de {f_fin.strftime('%B')} de {f_fin.year}")

    # Actualizar las sumas
    suma_dias = int(totalDias.cget("text")) + dias_incompletos
    suma_df = int(totalDFkwh.cget("text")) + consumo_incompleto_DF
    suma_d = int(totalDkwh.cget("text")) + consumoDF_incompleto

    totalDias.configure(text=suma_dias)
    totalDFkwh.configure(text=suma_df)
    totalDkwh.configure(text=suma_d)
    lbl28.configure(text=suma_d)

    # Guardar el nuevo periodo y resumen para la próxima ejecución
    ultimo_periodo_incompleto["desde"] = desde_incompleto
    ultimo_periodo_incompleto["hasta"] = hasta_incompleto

    ultimo_resumen_incompleto["dias"] = dias_incompletos
    ultimo_resumen_incompleto["df"] = consumo_incompleto_DF
    ultimo_resumen_incompleto["d"] = consumoDF_incompleto

def insertar_en_orden(tabla, nueva_fecha, valores):
    """Inserta una fila en orden cronológico dentro de la tabla"""
    filas = tabla.get_children()
    posicion = None

    # Convertir nueva fecha a número para comparación
    nueva_fecha = int(nueva_fecha)

    for fila in filas:
        fila_valores = tabla.item(fila, "values")
        fecha_existente = int(fila_valores[0])  # Primera columna es la fecha

        if nueva_fecha < fecha_existente:
            posicion = fila  # Insertar antes de esta fila
            break

    if posicion:
        tabla.insert(parent=tabla.parent(posicion), index=tabla.index(posicion), values=valores) # 🔹 Insertar en orden
    else:
        tabla.insert("", "end", values=valores)  # Si no encontró, insertar al final

def validar_fecha(treeview, periodo_desde, periodo_hasta):
    # Asegurar que las fechas sean objetos date
    if isinstance(periodo_desde, str):
        periodo_desde = datetime.strptime(periodo_desde, "%Y-%m-%d").date()
    if isinstance(periodo_hasta, str):
        periodo_hasta = datetime.strptime(periodo_hasta, "%Y-%m-%d").date()

    for item_id in treeview.get_children():
        item = treeview.item(item_id)
        valores = item["values"]

        if len(valores) >= 3:
            try:
                fecha_existente_desde = datetime.strptime(valores[1], "%Y-%m-%d").date()
                fecha_existente_hasta = datetime.strptime(valores[2], "%Y-%m-%d").date()
            except Exception as e:
                print(f"Error procesando fechas de fila: {valores}, error: {e}")
                continue

            # Validaciones de solapamiento
            if (
                fecha_existente_desde < periodo_desde < fecha_existente_hasta or
                fecha_existente_desde < periodo_hasta < fecha_existente_hasta or
                periodo_desde < fecha_existente_desde and periodo_hasta > fecha_existente_hasta
            ):
                messagebox.showerror(
                    "Error",
                    f"⚠ Las fechas seleccionadas ({periodo_desde} a {periodo_hasta}) se solapan con un período existente ({fecha_existente_desde} a {fecha_existente_hasta})."
                )
                return False

    return True  # Si pasó por todos sin solapamiento

def toggle_checkbox(event):
    item_id = Tabla.identify_row(event.y)
    column = Tabla.identify_column(event.x)

    if column == "#8" and item_id:  # "#8" es la columna SELECCION (índice 1-based)
        valores = list(Tabla.item(item_id, "values"))
        valores[7] = "✓" if valores[7] != "✓" else "✗"
        Tabla.item(item_id, values=valores)


agregar = ctk.CTkButton(calculos, text="Agregar",fg_color="#2b2b2b", bg_color="Gray", font=("Arial", int(fuente_base * 0.6), "bold"), hover_color="#FFC300", command=Agregar, state="disabled")
agregar.place(relx=0.4, rely=0.86, relwidth=0.2, relheight=0.06)
# Fin de Seccion de Calculos

# Seccion de Tabla de resultados
SeccionResultados = customtkinter.CTkFrame(ventana, fg_color="#C0C0C0", bg_color="#C0C0C0")
SeccionResultados.place(relx=0.35, rely=0.27, relwidth=0.65, relheight=0.75)

ResultadosTabla = customtkinter.CTkFrame(ventana, fg_color="Black", bg_color="Black")
ResultadosTabla.place(relx=0.43, rely=0.3, relwidth=0.4925, relheight=0.084)

def abrir_ventana_reporte():
    # Obtener el tamaño de la pantalla
    screen_width = ventana.winfo_screenwidth()
    screen_height = ventana.winfo_screenheight()

    ventana_reporte = customtkinter.CTkToplevel()
    ventana_reporte.title("Editor de Reporte")
    ventana_reporte.geometry(f"{screen_width}x{screen_height}")
    ventana_reporte.config(bg="#C0C0C0")

    estiloTablaTitulo = { 
        "fg_color": ("#C0C0C0"), 
        "bg_color": ("#C0C0C0"), 
        "text_color": ("White"),
        "font": ("Arial", int(fuente_base * 0.8),"bold")
    }

    estiloCamposTexto = { 
        "fg_color": ("White"), 
        "bg_color": ("White"), 
        "text_color": ("Black"),
        "font": ("Arial", int(fuente_base * 0.6),"bold")
    }

    titulo = customtkinter.CTkLabel(ventana_reporte, text="Vista Previa del Reporte", **estiloTablaTitulo)
    titulo.pack(pady=10)

    campos = {
        "Descripción": None,
        "Método": None,
        "Metodo2": None,
        "Giro": None,
        "Periodo": None,
        "Persona": None,
        "Situación": None
    }

    columnas = 2
    anchura_textbox = 600
    altura_textbox = 125
    separacion_horizontal = 0.5  # separa columnas
    separacion_vertical = 0.2   # separa filas

    for i, campo in enumerate(campos.keys()):
        fila = i // columnas
        columna = i % columnas

        # Posiciones relativas
        relx = 0.05 + columna * separacion_horizontal
        rely = 0.05 + fila * separacion_vertical

        # Etiqueta
        label = customtkinter.CTkLabel(ventana_reporte, text=campo + ":", anchor="w", **estiloTablaTitulo)
        label.place(relx=relx, rely=rely)

        # Caja de texto
        textbox = customtkinter.CTkTextbox(ventana_reporte, width=anchura_textbox, height=altura_textbox, **estiloCamposTexto)
        textbox.place(relx=relx, rely=rely + 0.05)

        campos[campo] = textbox

        # Switch para incluir la gráfica
        incluir_grafica_var = customtkinter.BooleanVar(value=False)

        switch_grafica = customtkinter.CTkSwitch(
            ventana_reporte,
            text="Incluir gráfica en el PDF",
            variable=incluir_grafica_var,
            onvalue=True,
            offvalue=False,
            switch_width=50,
            switch_height=25,
            font=("Arial", int(fuente_base * 0.6), "bold"),
            text_color="black",
            fg_color = "White",
            progress_color="Green",
            bg_color="#C0C0C0"
        )
        switch_grafica.place(relx=0.55, rely=0.9)  # Ajusta `rely` según tu diseño

    def confirmar_generacion():
        descripcion = campos["Descripción"].get("1.0", "end").strip()
        metodo = campos["Método"].get("1.0", "end").strip()
        metodo2 = campos["Metodo2"].get("1.0", "end").strip()
        giro = campos["Giro"].get("1.0", "end").strip()
        periodo = campos["Periodo"].get("1.0", "end").strip()
        persona = campos["Persona"].get("1.0", "end").strip()
        situacion = campos["Situación"].get("1.0", "end").strip()
      
        if not descripcion or not metodo or not metodo2 or not giro or not persona or not situacion:
            messagebox.showerror("Error", "Por favor completa todos los campos antes de generar el PDF.")
            return

        # Recolectar datos automáticos desde la GUI principal
        nombre = Nomrpu.cget("text")
        medidor = selected_medidor.get()
        tarifa = lbl25.cget("text")
        rpu = RPU
        anomalia_texto = anomalia.get() + " " + lbl33.cget("text")
        periodo_ajuste = periodo + " " + lbl31.cget("text")
        rpe = RPE
        nomrpe = Nomrpe.cget("text")
        tipo_tabla = "FACTOR"

        corriente_total = corriente.cget("text")
        voltaje_total = voltaje.cget("text")
        kvareal_total = kvareal.cget("text")
        kvacronometro_total = kvacronometro.cget("text")
        registracion_total = registracion.cget("text")
        Fact_total = FactAjus.cget("text")

        # Construir cpd_data directamente con encabezado y totales
        cpd_data = [["Corriente","Voltaje", "Kva Real", "KvA\nCronometro", "%\nRegistracion"]]  # encabezado

        # Agregar fila de totales
        cpd_data.append([
            str(corriente_total),
            str(voltaje_total),
            str(kvareal_total),
            str(kvacronometro_total),
            str(registracion_total)
        ])

        # Selección de archivo
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if not filepath:
            return  # Cancelado por el usuario
        
        print("🧪 Verificando cpd_data antes de enviar:")
        for i, fila in enumerate(cpd_data):
            print(f"Fila {i}: {fila} -> tipo: {type(fila)}, columnas: {len(fila)}")

        incluir_grafica = incluir_grafica_var.get()

        fechas_grafica, consumos_grafica, consumos_df_grafica = [], [], []

        if incluir_grafica:
            try:
                for item in Tabla.get_children():
                    fila = Tabla.item(item)["values"]
                    if len(fila) >= 6:
                        fechas_grafica.append(int(fila[0]))
                        consumos_grafica.append(float(fila[4]))
                        consumos_df_grafica.append(float(fila[5]))
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo extraer la información para la gráfica: {e}")
                return

        generar_reporte_pdf(
            nombre, direccion, medidor, tarifa, rpu, anomalia_texto,
            descripcion, metodo, giro, periodo_ajuste, persona, situacion,
            cpd_data, tipo_tabla, rpe, nomrpe, filepath,
            f"{float(Fact_total):.4f}", metodo2, incluir_grafica=incluir_grafica,
            fechas=fechas_grafica, consumos=consumos_grafica, consumos_df=consumos_df_grafica
        )

        ventana_reporte.destroy()
        messagebox.showinfo("Reporte generado", "📄 El PDF se ha guardado exitosamente")

    boton_generar = customtkinter.CTkButton(ventana_reporte, text="Generar", command=confirmar_generacion, fg_color="#e74c3c", hover_color="#cb4335", bg_color="#C0C0C0", font=("Arial", int(fuente_base * 0.55), "bold"))
    boton_generar.place(relx=0.45, rely=0.9, relwidth=0.08, relheight=0.04)

def exportar_a_excel():
    try:
        # Seleccionar plantilla
        plantilla_path = "Plantilla Factor Ajuste.xlsx"  # Asegúrate que está en el mismo directorio o usa path absoluto
        wb = load_workbook(plantilla_path)
        ws = wb["X F. A."]

        # Datos principales
        nombre = Nomrpu.cget("text")
        agen = agencia.cget("text")
        medidor = selected_medidor.get()
        tarifa = lbl25.cget("text")
        rpu = RPU
        cuent = cuenta.cget("text")
        anomalia = lbl33.cget("text")
        periodo_ajuste = lbl31.cget("text")
        nomrpe = Nomrpe.cget("text")
        corr = corriente.cget("text")
        volt = voltaje.cget("text")
        kvre = kvareal.cget("text")
        kvaCro = kvacronometro.cget("text")
        reg = registracion.cget("text")
        fa = FactAjus.cget("text")

        # Insertar en celdas específicas (ajusta si tu archivo es diferente)
        ws["F10"] = nombre
        ws["C9"] = rpu
        ws["F9"] = medidor
        ws["K5"] = datetime.now().strftime("%A, %d %B, %Y")
        ws["J10"] = agen
        ws["C10"] = cuent
        ws["E16"] = corr
        ws["F16"] = volt
        ws["G16"] = kvre
        ws["H16"] = kvaCro
        ws["I16"] = reg
        ws["H18"] = fa

        # Eliminar datos anteriores (opcional)
        fila_inicio = 22
        for row in ws.iter_rows(min_row=fila_inicio, max_row=fila_inicio + 100, min_col=5, max_col=11):
            for cell in row:
                cell.value = None

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Copiar valores de la tabla sin copiar estilos
        for i, item_id in enumerate(Tabla.get_children()):
            valores = Tabla.item(item_id)["values"]
            for j, valor in enumerate(valores[:7]):  # Solo los primeros 7 valores
                col = 5 + j  # Comienza en la columna E (5)
                cell = ws.cell(row=fila_inicio + i, column=col)
                cell.value = valor
                cell.border = thin_border  # 👈 Aquí aplicamos el borde
        
        # Insertar totales al final de la tabla
        fila_total = fila_inicio + len(Tabla.get_children())
        relleno_totales = PatternFill(fill_type="solid", fgColor="FFFF00")  # Amarillo
        fuente_totales = Font(name="Arial", size=24, bold=True)  # Arial, 24pt, negritas


        ws[f"G{fila_total}"] = "TOTAL"
        ws[f"H{fila_total}"] = totalDias.cget("text")
        ws[f"I{fila_total}"] = totalCFRkwh.cget("text")
        ws[f"J{fila_total}"] = totalDFkwh.cget("text")
        ws[f"K{fila_total}"] = totalDkwh.cget("text")

        for col in range(5, 12):  # Columnas E (5) a K (11)
            cell = ws.cell(row=fila_total, column=col)
            cell.border = thin_border
            cell.fill = relleno_totales
            cell.font = fuente_totales
        
        # Insertar valores inferiores dinámicamente
        fila_base = fila_inicio + len(Tabla.get_children()) + 2

        ws[f"F{fila_base}"] = "Tarifa del servicio"
        ws[f"F{fila_base}"].alignment = Alignment(horizontal="left")
        ws[f"H{fila_base}"] = tarifa
        ws[f"H{fila_base}"].alignment = Alignment(horizontal="center")
        ws[f"I{fila_base}"] = "Suministro en BAJA-BAJA"
        ws[f"I{fila_base}"].alignment = Alignment(horizontal="left")

        ws[f"F{fila_base + 2}"] = "Recuperación"
        ws[f"F{fila_base}"].alignment = Alignment(horizontal="left")
        ws[f"H{fila_base + 2}"] = totalCFRkwh.cget("text")
        ws[f"H{fila_base}"].alignment = Alignment(horizontal="center")
        ws[f"I{fila_base + 2}"] = "KWH TOTALES"
        ws[f"I{fila_base}"].alignment = Alignment(horizontal="left")

        ws[f"F{fila_base + 4}"] = "Periodo del ajuste"
        ws[f"F{fila_base}"].alignment = Alignment(horizontal="left")
        ws[f"I{fila_base + 4}"] = periodo_ajuste

        ws[f"G{fila_base + 8}"] = f"Anomalía: {anomalia}"

        ws[f"H{fila_base + 15}"] = f"___________________________________________________________________________"
        ws[f"H{fila_base + 16}"] = f"Ing. {nomrpe}"
        ws[f"H{fila_base + 17}"] = f"Oficina de medición Tula"

        # Guardar archivo
        ruta_salida = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar reporte de Factor de Ajuste"
        )
        if ruta_salida:
            wb.save(ruta_salida)
            messagebox.showinfo("Exportación Exitosa", "📄 El archivo Excel se ha guardado correctamente.")
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al generar el Excel:\n{e}")
        
estiloTablaTitulo = { 
        "fg_color": ("#2fd134"), 
        "bg_color": ("#2fd134"), 
        "text_color": ("White"), 
        "width": 137, 
        "height": 38, 
        "font": ("Arial", int(fuente_base * 0.65),"bold")
}

estiloTablaDatos = { 
        "fg_color": ("White"), 
        "bg_color": ("White"), 
        "text_color": ("Black"), 
        "width": 137, 
        "height": 30, 
        "font": ("Arial", int(fuente_base * 0.65))
}

estiloTablaFinal = { 
        "fg_color": ("White"), 
        "bg_color": ("White"), 
        "text_color": ("Black"), 
        "width": 120, 
        "height": 40, 
        "font": ("Arial", int(fuente_base * 0.65), "bold")
}

# Configurar columnas y filas para que se expandan
for i in range(1, 7):  # columnas 1 a 6
    ResultadosTabla.columnconfigure(i, weight=1)
for i in range(1, 3):  # filas 1 y 2
    ResultadosTabla.rowconfigure(i, weight=1)

col1 = customtkinter.CTkLabel(ResultadosTabla, text="Corriente", **estiloTablaTitulo)
col1.grid(row=1, column=1, padx=1, pady=1, sticky="nsew")

corriente = customtkinter.CTkLabel(ResultadosTabla, text="0.00", **estiloTablaDatos)
corriente.grid(row=2, column=1, padx=1, pady=1, sticky="nsew")

col2 = customtkinter.CTkLabel(ResultadosTabla, text="Voltaje", **estiloTablaTitulo)
col2.grid(row=1, column=2, padx=1, pady=1, sticky="nsew")

voltaje = customtkinter.CTkLabel(ResultadosTabla, text="0.00", **estiloTablaDatos)
voltaje.grid(row=2, column=2, padx=1, pady=1, sticky="nsew")

col3 = customtkinter.CTkLabel(ResultadosTabla, text="kVA real", **estiloTablaTitulo)
col3.grid(row=1, column=3, padx=1, pady=1, sticky="nsew")

kvareal = customtkinter.CTkLabel(ResultadosTabla, text="0.00", **estiloTablaDatos)
kvareal.grid(row=2, column=3, padx=1, pady=1, sticky="nsew")

col4 = customtkinter.CTkLabel(ResultadosTabla, text="kVA Cronometro", **estiloTablaTitulo)
col4.grid(row=1, column=4, padx=1, pady=1, sticky="nsew")

kvacronometro = customtkinter.CTkLabel(ResultadosTabla, text="0.00", **estiloTablaDatos)
kvacronometro.grid(row=2, column=4, padx=1, pady=1, sticky="nsew")

col5 = customtkinter.CTkLabel(ResultadosTabla, text="% Registración", **estiloTablaTitulo)
col5.grid(row=1, column=5, padx=1, pady=1, sticky="nsew")

registracion = customtkinter.CTkLabel(ResultadosTabla, text="0.00", **estiloTablaDatos)
registracion.grid(row=2, column=5, padx=1, pady=1, sticky="nsew")

col6 = customtkinter.CTkLabel(ResultadosTabla, text="FACTOR DE\n AJUSTE", **estiloTablaTitulo)
col6.grid(row=1, column=6, padx=1, pady=1, sticky="nsew")

FactAjus = customtkinter.CTkLabel(ResultadosTabla, text="0.00", **estiloTablaDatos)
FactAjus.grid(row=2, column=6, padx=1, pady=1, sticky="nsew")

TablaFinal = customtkinter.CTkFrame(ventana, fg_color="Black", bg_color="Black")
TablaFinal.place(relx=0.395, rely=0.4, relwidth=0.564, relheight=0.096)

for i in range(1, 8):  # 7 columnas
    TablaFinal.columnconfigure(i, weight=1)
for i in range(1, 3):  # 2 filas
    TablaFinal.rowconfigure(i, weight=1)

lbl15 = customtkinter.CTkLabel(TablaFinal, text="Consumos facturados\n reflejados en SICOM", fg_color="#fff301", bg_color="#fff301", text_color="Black", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl15.grid(row=1, column=1, columnspan=4, padx=1, pady=1, sticky="nsew")

lbl16 = customtkinter.CTkLabel(TablaFinal, text="Debio\n facturar", fg_color="Orange", bg_color="Orange", text_color="Black", font= ("Arial", int(fuente_base * 0.65),"bold"))
lbl16.grid(row=1, column=5, padx=1, pady=1, sticky="nsew")

lbl17 = customtkinter.CTkLabel(TablaFinal, text="Diferencia", width=120, height=40, fg_color="#ff0000", bg_color="#ff0000", text_color="Black", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl17.grid(row=1, column=6, padx=1, pady=1, sticky="nsew")

lbl35 = customtkinter.CTkLabel(TablaFinal, text="Seleccion", width=120, height=40, fg_color="Gray", bg_color="Gray", text_color="Black", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl35.grid(row=1, column=7, padx=1, pady=1, sticky="nsew")

lbl18 = customtkinter.CTkLabel(TablaFinal, text="Fecha", **estiloTablaFinal)
lbl18.grid(row=2, column=1, padx=1, pady=1, sticky="nsew")

lbl19 = customtkinter.CTkLabel(TablaFinal, text="Fecha \n Desde - Hasta", width=223, height=40, fg_color="White", bg_color="White", text_color="Black", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl19.grid(row=2, column=2, padx=1, pady=1, sticky="nsew")

lbl20 = customtkinter.CTkLabel(TablaFinal, text="Dias", **estiloTablaFinal)
lbl20.grid(row=2, column=3, padx=1, pady=1, sticky="nsew")

lbl21 = customtkinter.CTkLabel(TablaFinal, text="kWh Total", **estiloTablaFinal)
lbl21.grid(row=2, column=4, padx=1, pady=1, sticky="nsew")

lbl22 = customtkinter.CTkLabel(TablaFinal, text="kWh Total", **estiloTablaFinal)
lbl22.grid(row=2, column=5, padx=1, pady=1, sticky="nsew")

lbl23 = customtkinter.CTkLabel(TablaFinal, text="kWh Total",  **estiloTablaFinal)
lbl23.grid(row=2, column=6, padx=1, pady=1, sticky="nsew")

lbl36 = customtkinter.CTkLabel(TablaFinal, text="✓",  **estiloTablaFinal)
lbl36.grid(row=2, column=7, padx=1, pady=1, sticky="nsew")

DatosBD = customtkinter.CTkFrame(ventana, fg_color="Black", bg_color="Black")
DatosBD.pack_propagate(False)
DatosBD.place(relx=0.395, rely=0.5, relwidth=0.573, relheight=0.3)

# Apartado para los datos de la BD

style = ttk.Style()
style.configure("Resultados.Treeview", font=("Arial", int(fuente_base * 0.8)), rowheight=40)

Tabla = ttk.Treeview(DatosBD, style="Resultados.Treeview", columns=("FECHA", "DESDE", "HASTA", "DIAS", "CONSUMO", "CONSUMNO_DF", "CONSUMO_D", "SELECCION"), show="tree")
Tabla.column("#0", width=0, stretch=tk.NO)
Tabla.column("FECHA", anchor=CENTER, width=25)
Tabla.column("DESDE", anchor=CENTER, width=8)
Tabla.column("HASTA", anchor=CENTER, width=8)
Tabla.column("DIAS", anchor=CENTER, width=25)
Tabla.column("CONSUMO", anchor=CENTER, width=25)
Tabla.column("CONSUMNO_DF", anchor=CENTER, width=25)
Tabla.column("CONSUMO_D", anchor=CENTER, width=25)
Tabla.column("SELECCION", anchor=CENTER, width=25)

Tabla.bind("<Button-1>", toggle_checkbox)

for col in ("FECHA", "DESDE", "HASTA", "DIAS", "CONSUMO", "CONSUMNO_DF", "CONSUMO_D", "SELECCION"):
    Tabla.column(col, anchor=tk.CENTER, width=115)
    Tabla.heading(col, text="") 

# Crear Scrollbar
scrollbar = ttk.Scrollbar(DatosBD, orient="vertical", command=Tabla.yview)
Tabla.configure(yscrollcommand=scrollbar.set)

# Posicionar elementos
Tabla.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

TotalBD = customtkinter.CTkFrame(ventana, width=850, height=40, fg_color="Black", bg_color="Black")
TotalBD.place(relx=0.395, rely=0.8, relwidth=0.563, relheight=0.048)

for i in range(1, 8):  # 7 columnas
    TotalBD.columnconfigure(i, weight=1)
for i in range(1, 3):  # 2 filas
    TotalBD.rowconfigure(i, weight=1)

lblTotal = customtkinter.CTkLabel(TotalBD, text="Total :",  width=344, height=40, fg_color="White", bg_color="White", text_color="Black", font=("Arial", int(fuente_base * 0.65),"bold"))
lblTotal.grid(row=1, column=1, padx=1, pady=1, sticky="nsew")

totalDias = customtkinter.CTkLabel(TotalBD, text="00.0",  **estiloTablaFinal)
totalDias.grid(row=1, column=3, padx=1, pady=1, sticky="nsew")

totalCFRkwh = customtkinter.CTkLabel(TotalBD, text="00.0",  **estiloTablaFinal)
totalCFRkwh.grid(row=1, column=4, padx=1, pady=1, sticky="nsew")

totalDFkwh = customtkinter.CTkLabel(TotalBD, text="00.0",  **estiloTablaFinal)
totalDFkwh.grid(row=1, column=5, padx=1, pady=1, sticky="nsew")

totalDkwh = customtkinter.CTkLabel(TotalBD, text="00.0",  **estiloTablaFinal)
totalDkwh.grid(row=1, column=6, padx=1, pady=1, sticky="nsew")

btnConfirmar = ctk.CTkButton(TotalBD, text="Confirmar",fg_color="White", bg_color="Gray", width=120, height=40, text_color="Black", font=("Arial", 14, "bold"), hover_color="Gray", command=confirmar_calculo, state="disabled")
btnConfirmar.grid(row=1, column=7, padx=1, pady=1, sticky="nsew")

DatosFinales = { 
        "fg_color": ("#C0C0C0"), 
        "bg_color": ("#C0C0C0"), 
        "text_color": ("Black"), 
        "width": 80, 
        "height": 40, 
        "font": ("Arial", int(fuente_base * 0.65), "bold")
}

lbl24 = customtkinter.CTkLabel(ventana, text="Tarifa del servicio :", **DatosFinales)
lbl24.place(relx=0.5, rely=0.85)

lbl25 = customtkinter.CTkLabel(ventana, text="00", **DatosFinales)
lbl25.place(relx=0.585, rely=0.85)

lbl26 = customtkinter.CTkLabel(ventana, text="Suministro en Baja - Baja", **DatosFinales)
lbl26.place(relx=0.63, rely=0.85)

lbl27 = customtkinter.CTkLabel(ventana, text="Recuperacion :", **DatosFinales)
lbl27.place(relx=0.5, rely=0.88)

lbl28 = customtkinter.CTkLabel(ventana, text="0000", **DatosFinales)
lbl28.place(relx=0.58, rely=0.88)

lbl29 = customtkinter.CTkLabel(ventana, text="kWh Totales", **DatosFinales)
lbl29.place(relx=0.63, rely=0.88)

lbl30 = customtkinter.CTkLabel(ventana, text="Periodo del ajuste :", **DatosFinales)
lbl30.place(relx=0.5, rely=0.915)

lbl31 = customtkinter.CTkLabel(ventana, text="", **DatosFinales)
lbl31.place(relx=0.59, rely=0.915)

lbl32 = customtkinter.CTkLabel(ventana, text="Anomalia :", **DatosFinales)
lbl32.place(relx=0.5, rely=0.95)

lbl33 = customtkinter.CTkLabel(ventana, text="-------------------", **DatosFinales)
lbl33.place(relx=0.59, rely=0.95)

boton_excel = customtkinter.CTkButton(ventana, text="Exportar a Excel", command=exportar_a_excel, fg_color="#4CAF50", hover_color="#45a049", bg_color="#C0C0C0", font=("Arial", int(fuente_base * 0.55), "bold"))
boton_excel.place(relx=0.86, rely=0.88, relwidth=0.08, relheight=0.04)  # Ajusta posición según diseño

btn_pdf = customtkinter.CTkButton(ventana, text="Generar PDF", command=abrir_ventana_reporte, fg_color="#e74c3c", hover_color="#cb4335", bg_color="#C0C0C0", font=("Arial", int(fuente_base * 0.55), "bold"))
btn_pdf.place(relx=0.86, rely=0.94, relwidth=0.08, relheight=0.04)  # Puedes ajustar relx para moverlo más a la derecha o izquierda
  
ventana.mainloop()