from tkinter import *
from tkinter import messagebox
from PIL import Image, ImageTk
import customtkinter
import customtkinter as ctk
from tkinter import ttk
import tkinter as tk
from tkcalendar import DateEntry
from datetime import datetime, timedelta
import locale
from Database.Conexion_SQL_Server import *
from Database.Datos import * 
import sys
import math
from Generacion_Reportes import generar_reporte_pdf
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from tkinter import filedialog
from openpyxl import load_workbook
from copy import copy

ventana = customtkinter.CTk()
ventana.title("Carga Instantanea")
ventana.config(bg="#C0C0C0")
customtkinter.set_appearance_mode("dark")

# Obtener el tamaño de la pantalla
screen_width = ventana.winfo_screenwidth()
screen_height = ventana.winfo_screenheight()

# Configurar la ventana para que ocupe todo el tamaño de la pantalla
ventana.geometry(f"{screen_width}x{screen_height}")

# Escalar fuente
fuente_base = int(screen_width * 0.015)

# Poner los elementos del programa en Español
locale.setlocale(locale.LC_TIME, "es_MX")

# sys.argv contiene los argumentos pasados
# RPE = sys.argv[1]  # Primer argumento
# RPU = sys.argv[2]    # Segundo argumento
RPU = 272020100257
RPE = "9L99B"

# 🔹 Cargar imagen correctamente con PIL
imagen = Image.open("Imagenes/Logo_CFE.png").resize((int(screen_width*0.3), int(screen_height*0.1))).convert("RGBA")
imagen_tk = ImageTk.PhotoImage(imagen)

imagen2 = Image.open("Imagenes/Logo_Medicion.jpg").resize((int(screen_width*0.15), int(screen_height*0.16))).convert("RGBA")
imagen_tk2 = ImageTk.PhotoImage(imagen2)

# 🔹 Encabezado
encabezado = customtkinter.CTkFrame(ventana, fg_color="White", bg_color="Black")
encabezado.place(relx=0, rely=0, relwidth=1, relheight=0.2)

img = customtkinter.CTkLabel(encabezado, image=imagen_tk, text=" ")
img.place(relx=0.025, rely=0.15)
img.image = imagen_tk

titulo = customtkinter.CTkLabel(encabezado, text="COMISION FEDERAL DE ELECTRICIDAD \n DIVISION CENTRO ORIENTE \n ZONA DE DISTRIBUCION TULA",
                                    bg_color="White", fg_color="White", text_color="Black", font=("Arial", int(fuente_base * 0.9), "bold"))
titulo.place(relx=0.39, rely=0.2)

img2 = customtkinter.CTkLabel(encabezado, image=imagen_tk2, text=" ")
img2.place(relx=0.825, rely=0.08)
img2.image = imagen_tk2
# Fin de Encabezado

# Datos Generales
# Datos RPE
ListaRPE = {
"9L99B": "Francisco Javier Velazquez Cervantes",
"JA117": "Jerzael Abisai Mera Zapatero"
}

datos = customtkinter.CTkFrame(ventana, width=screen_width, height=110, fg_color="green", bg_color="Black")
datos.place(relx=0, rely=0.15, relwidth=1, relheight=0.12)

estiloDatos = { 
        "fg_color": ("Green"), 
        "bg_color": ("Green"), 
        "text_color": ("White"), 
        "width": 50, 
        "height": 20, 
        "font": ("Arial", int(fuente_base * 0.65),"bold")
}

def ExtraccionCuenta():
    texto = cuenta.cget("text")  # Obtener el texto del Label

    Cclo = texto[:2]        # Primeras 2 letras
    pblacion = texto[7:9]   # Caracteres 6 y 7
    rta = texto[9:12]       # 3 caracteres después
    flio = texto[12:16]       # Últimos 4 caracteres

    ciclo.configure(text=Cclo)
    poblacion.configure(text=pblacion)
    ruta.configure(text=rta)
    folio.configure(text=flio)

ahora = datetime.now()
fecha_hora = ahora.strftime("%A, %d %B, %Y")

lbl1 = customtkinter.CTkLabel(datos, text="R.P.E :", **estiloDatos)
lbl1.place(relx=0.05, rely=0.2)

rpe = customtkinter.CTkLabel(datos, text=RPE, **estiloDatos)
rpe.place(relx=0.1, rely=0.2)

lbl2 = customtkinter.CTkLabel(datos, text="Nombre :", **estiloDatos)
lbl2.place(relx=0.25, rely=0.2)

Nomrpe = customtkinter.CTkLabel(datos, text="Nombre del RPE", **estiloDatos)
Nomrpe.place(relx=0.3, rely=0.2)

FechaAct = customtkinter.CTkLabel(datos, text=fecha_hora, **estiloDatos)
FechaAct.place(relx=0.84, rely=0.2)

# Buscar el RPE dentro del diccionario de trabajadores
nombre_rpe = ListaRPE.get(RPE, "Trabajador no encontrado")  # Si no está en el diccionario, muestra "Trabajador no encontrado"
    
Nomrpe.configure(text=nombre_rpe)

# Datos RPU - Fila 1
DatosRPU = BD.datosRPU(RPU)
MedidoresRPU = BD.medidores(RPU)

lbl3 = customtkinter.CTkLabel(datos, text="R.P.U :", **estiloDatos)
lbl3.place(relx=0.05, rely=0.4)

rpu = customtkinter.CTkLabel(datos, text=RPU, **estiloDatos)
rpu.place(relx=0.1, rely=0.4)

lbl4 = customtkinter.CTkLabel(datos, text="Cuenta :", **estiloDatos)
lbl4.place(relx=0.25, rely=0.4)

cuenta = customtkinter.CTkLabel(datos, text="23DV13B162333130", **estiloDatos)
cuenta.place(relx=0.3, rely=0.4)

lbl5 = customtkinter.CTkLabel(datos, text="Nombre :", **estiloDatos)
lbl5.place(relx=0.45, rely=0.4)

Nomrpu = customtkinter.CTkLabel(datos, text="Nombre del RPU", **estiloDatos)
Nomrpu.place(relx=0.5, rely=0.4)

lbl6 = customtkinter.CTkLabel(datos, text="Medidor :", **estiloDatos)
lbl6.place(relx=0.65, rely=0.4)

medidores = [registro[1] for registro in MedidoresRPU]
selected_medidor = ctk.StringVar()
selected_medidor.set(medidores[0] if medidores else "")

medidor = customtkinter.CTkOptionMenu(datos, variable=selected_medidor ,values=medidores, **estiloDatos)
medidor.place(relx=0.7, rely=0.4)

# Datos RPU - Fila 2
direccion = ""

lbl7 = customtkinter.CTkLabel(datos, text="Ciclo :",  **estiloDatos)
lbl7.place(relx=0.05, rely=0.6)

ciclo = customtkinter.CTkLabel(datos, text=" ", **estiloDatos)
ciclo.place(relx=0.1, rely=0.6)

lbl8 = customtkinter.CTkLabel(datos, text="Poblacion :", **estiloDatos)
lbl8.place(relx=0.25, rely=0.6)

poblacion = customtkinter.CTkLabel(datos, text=" ", **estiloDatos)
poblacion.place(relx=0.3, rely=0.6)

lbl9 = customtkinter.CTkLabel(datos, text="Ruta :", **estiloDatos)
lbl9.place(relx=0.45, rely=0.6)

ruta = customtkinter.CTkLabel(datos, text=" ", **estiloDatos)
ruta.place(relx=0.5, rely=0.6)

lbl10 = customtkinter.CTkLabel(datos, text="Folio :", **estiloDatos)
lbl10.place(relx=0.65, rely=0.6)

folio = customtkinter.CTkLabel(datos, text=" ", **estiloDatos)
folio.place(relx=0.7, rely=0.6)

lbl11 = customtkinter.CTkLabel(datos, text="Agencia :", **estiloDatos)
lbl11.place(relx=0.85, rely=0.6)

agencia = customtkinter.CTkLabel(datos, text="Nombre de la Agencia", **estiloDatos)
agencia.place(relx=0.9, rely=0.6)

# Diccionario de las agencias
AGENCIAS = {
    "A": "Progreso",
    "B": "Tula",
    "C": "Tlaxcoapan",
    "D": "Tepeji"
}

# Cargar Datos de RPU de la BD en las etiquetas

if DatosRPU and len(DatosRPU) >= 4:
    Nomrpu.configure(text=DatosRPU[1])  # Nombre RPU
    direccion = DatosRPU[2]  # Nombre RPU
    cuenta.configure(text=DatosRPU[3])  # Cuenta

     # Obtener la agencia del resultado
    codigo_agencia = DatosRPU[4]
    
    # Convertir el código a su nombre correspondiente
    nombre_agencia = AGENCIAS.get(codigo_agencia, "Desconocida")  # Si no está en el diccionario, muestra "Desconocida"
    
    agencia.configure(text=nombre_agencia)
else:
    # Si no hay datos, mostrar un mensaje de "No encontrado"
    Nomrpu.configure(text="No encontrado")
    direccion = None
    cuenta.configure(text="No encontrado")
    agencia.configure(text="No encontrado")

# Fin de Datos Generales

# Llamar a la función inmediatamente al inicio
ExtraccionCuenta()

# Seccion de Calculos
calculos = customtkinter.CTkFrame(ventana, fg_color="gray", bg_color="Black")
calculos.place(relx=0, rely=0.27, relwidth=0.35, relheight=0.75)

lbl12 = customtkinter.CTkLabel(calculos, text="BASE DE CALCULO PARA REALIZAR AJUSTE", fg_color="Gray", bg_color="Gray", text_color="White", width=75, height=20, font=("Arial", 16,"bold"))
lbl12.place(relx=0.22, rely=0.02)

# Tabla
class TablaApp(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color="#2b2b2b", bg_color="#2b2b2b")
        self.place(relx=0.02, rely=0.07, relwidth=0.96, relheight=0.2)

         # Usar grid en lugar de pack para mejor control
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.canvas = ctk.CTkCanvas(self, bg="#2b2b2b", highlightthickness=0)
        self.canvas.grid(row=0, column=0, sticky="nsew")

        self.scrollbar = ctk.CTkScrollbar(self, command=self.canvas.yview, fg_color="#2b2b2b")
        self.scrollbar.grid(row=0, column=1, sticky="ns")

        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.table_frame = ctk.CTkFrame(self.canvas, fg_color="#2b2b2b", bg_color="#2b2b2b")
        self.canvas_window = self.canvas.create_window((0, 0), window=self.table_frame, anchor="nw")

        self.table_frame.bind("<Configure>", self.on_frame_configure)
        self.canvas.bind("<Configure>", self.on_canvas_configure)

        self.table = []
        self.create_table()

        self.add_button = ctk.CTkButton(parent, text="Agregar fila", command=self.add_row, bg_color="Gray", fg_color="Green", hover_color="#2fd134")
        self.add_button.place(relx=0.2, rely=0.3, relwidth=0.25, relheight=0.05)

        self.remove_button = ctk.CTkButton(parent, text="Quitar fila", command=self.remove_row, bg_color="Gray", fg_color="Darkred", hover_color="Red")
        self.remove_button.place(relx=0.5, rely=0.3, relwidth=0.25, relheight=0.05)

        self.table_frame.bind("<Configure>", self.on_frame_configure)

    def create_table(self):
        headers = ["Fase", "Corriente", "Voltaje", "Potencia", "Horas/Uso", "C.P.D"]
        for col, header in enumerate(headers):
            # Establece un tamaño mínimo y permite que la columna se expanda
            self.table_frame.columnconfigure(col, weight=1, minsize=100)

            label = ctk.CTkLabel(
                self.table_frame,
                text=header,
                fg_color="green",
                bg_color="green",
                text_color="white",
                font=("Arial", int(fuente_base * 0.55), "bold")
            )
            label.grid(row=0, column=col, sticky="nsew", padx=5, pady=5)

        for _ in range(1):  # Agrega 1 fila inicial
            self.add_row()

    def add_row(self):
        row = []
        row_index = len(self.table) + 1
        self.table_frame.rowconfigure(row_index, weight=1, minsize=40)

        for j in range(6):
            if j in [3, 5]:  # Potencia y C.P.D como labels
                label = ctk.CTkLabel(
                    self.table_frame,
                    text="0.00",
                    fg_color="White",
                    text_color="Black"
                )
                label.grid(row=row_index, column=j, sticky="nsew", padx=5, pady=5)
                row.append(label)
            else:
                var = ctk.StringVar()
                entry = ctk.CTkEntry(
                    self.table_frame,
                    fg_color="White",
                    bg_color="White",
                    border_color="White",
                    text_color="Black",
                    textvariable=var
                )
                entry.grid(row=row_index, column=j, sticky="nsew", padx=5, pady=5)
                var.trace_add("write", lambda *args, row=row: self.update_labels(row))
                row.append(entry)

        self.table.append(row)
        self.update_scrollregion()

    def update_labels(self, row):
        """Función que calcula Potencia y CPD en base a los valores ingresados en Corriente, Voltaje y Horas/Uso."""
        try:
            corriente = float(row[1].get()) if row[1].get() else 0
            voltaje = float(row[2].get()) if row[2].get() else 0
            hrs_uso = float(row[4].get()) if row[4].get() else 0

            potencia = ((corriente * voltaje)/1000)*0.9 # Cálculo de Potencia
            cpd = potencia * hrs_uso  # Cálculo de C.P.D

            row[3].configure(text=f"{potencia:.3f}")  # Actualiza el Label de Potencia
            row[5].configure(text=f"{cpd:.4f}")  # Actualiza el Label de CPD
        except ValueError:
            pass  # Ignorar errores en caso de valores inválidos

    def remove_row(self):
        if self.table:
            row = self.table.pop()
            for widget in row:
                widget.destroy()
            self.update_scrollregion()

    def on_frame_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def update_scrollregion(self):
        self.canvas.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def on_canvas_configure(self, event):
        canvas_width = event.width
        self.canvas.itemconfig(self.canvas_window, width=canvas_width)

    def obtener_datos_tabla(self):
        datos = []
        for fila in self.table:
            fila_datos = []
            for idx, widget in enumerate(fila):
                if isinstance(widget, ctk.CTkEntry):
                    fila_datos.append(widget.get())
                elif isinstance(widget, ctk.CTkLabel):
                    fila_datos.append(widget.cget("text"))
            datos.append(fila_datos)
        return datos

tabla = TablaApp(calculos)
# Fin de la tabla

# Apartado del Periodo
lbl13 = customtkinter.CTkLabel(calculos, text="Periodo", fg_color="Gray", bg_color="Gray", text_color="White", width=75, height=20, font=("Arial", 16,"bold"))
lbl13.place(relx=0.05, rely=0.35)

class PeriodoSelector(ctk.CTkFrame):
    def __init__(self, master=None):
        super().__init__(master)
        self.place(relwidth=0.9, relheight=0.12)
        self.pack_propagate(False)

        # Label Desde
        self.label_desde = ctk.CTkLabel(self, text="Desde (Año-Mes-Día):")
        self.label_desde.place(relx=0.2, rely=0.1)

        # Selector de fecha "Desde" (Formato completo)
        self.date_desde = DateEntry(self, background='darkgreen', 
                                    foreground='white', borderwidth=2, year=2025, 
                                    date_pattern="yyyy-mm-dd", font=(int(fuente_base * 0.35)))  
        self.date_desde.place(relx=0.23, rely=0.5, relwidth=0.22, relheight=0.25)

        # Label Hasta
        self.label_hasta = ctk.CTkLabel(self, text="Hasta (Año-Mes-Día):")
        self.label_hasta.place(relx=0.55, rely=0.1)

        # Selector de fecha "Hasta" (Formato completo)
        self.date_hasta = DateEntry(self, background='darkgreen', 
                                    foreground='white', borderwidth=2, year=2025, 
                                    date_pattern="yyyy-mm-dd", font=(int(fuente_base * 0.45)))  
        self.date_hasta.place(relx=0.57, rely=0.5, relwidth=0.22, relheight=0.25)

    def get_range(self):
        # Obtener la fecha seleccionada y mostrarla con Año-Mes-Día
        desde_año = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%Y")
        desde_mes = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%m")
        desde_mes_palabra = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%B")
        desde_dia = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%d")

        hasta_año = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%Y")
        hasta_mes = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%m")
        hasta_mes_palabra = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%B")
        hasta_dia = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%d")
        
        return desde_año, desde_mes, desde_mes_palabra, desde_dia, hasta_año, hasta_mes, hasta_mes_palabra, hasta_dia


periodo_selector = PeriodoSelector(calculos)
periodo_selector.place(relx=0.05, rely=0.39)  # Ajusta la posición
# Fin del Apartado Periodo

lbl14 = customtkinter.CTkLabel(calculos, text="Anomalia :", fg_color="Gray", bg_color="Gray", text_color="White", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl14.place(relx=0.05, rely=0.52)

anomalia = customtkinter.CTkEntry(calculos, fg_color="White", bg_color="Gray", text_color="Black")
anomalia.place(relx=0.2, rely=0.52, relwidth=0.2, relheight=0.05)

class PeriodoIncompleto(ctk.CTkFrame):
    def __init__(self, master=None, periodo_selector=None):
        super().__init__(master)
        self.place(relwidth=0.9, relheight=0.12)
        self.pack_propagate(False)

        self.periodo_selector = periodo_selector  # Guardamos la referencia del PeriodoSelector

        # Label Desde
        self.label_desde = ctk.CTkLabel(self, text="Desde (Año-Mes-Día):")
        self.label_desde.place(relx=0.2, rely=0.1)

        # Selector de fecha "Desde" (Formato completo)
        self.date_desde = DateEntry(self, background='darkred', 
                                    foreground='white', borderwidth=2, year=2025, 
                                    date_pattern="yyyy-mm-dd", font=(int(fuente_base * 0.35)))  
        self.date_desde.place(relx=0.23, rely=0.5, relwidth=0.22, relheight=0.25)

        # Label Hasta
        self.label_hasta = ctk.CTkLabel(self, text="Hasta (Año-Mes-Día):")
        self.label_hasta.place(relx=0.55, rely=0.1)

        # Selector de fecha "Hasta" (Formato completo)
        self.date_hasta = DateEntry(self, background='darkred', 
                                    foreground='white', borderwidth=2, year=2025, 
                                    date_pattern="yyyy-mm-dd", font=(int(fuente_base * 0.35)))  
        self.date_hasta.place(relx=0.57, rely=0.5, relwidth=0.22, relheight=0.25)

    def get_range(self):
       
        # Obtener la fecha seleccionada y mostrarla con Año-Mes-Día
        incompleto_desde_año = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%Y")
        incompleto_desde_mes = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%m")
        incompleto_desde_mes_palabra = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%B")
        incompleto_desde_dia = datetime.strptime(self.date_desde.get(), "%Y-%m-%d").strftime("%d")

        incompleto_hasta_año = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%Y")
        incompleto_hasta_año_formato = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%Y")[-2:]
        incompleto_hasta_mes = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%m")
        incompleto_hasta_mes_palabra = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%B")
        incompleto_hasta_dia = datetime.strptime(self.date_hasta.get(), "%Y-%m-%d").strftime("%d")
        
        return incompleto_desde_año, incompleto_desde_mes, incompleto_desde_mes_palabra, incompleto_desde_dia, incompleto_hasta_año, incompleto_hasta_año_formato, incompleto_hasta_mes, incompleto_hasta_mes_palabra, incompleto_hasta_dia


lbl34 = customtkinter.CTkLabel(calculos, text="Periodo Incompleto", fg_color="Gray", bg_color="Gray", text_color="White", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl34.place(relx=0.05, rely=0.59)

periodo_selector_incompleto = PeriodoIncompleto(calculos, periodo_selector)
periodo_selector_incompleto.place(relx=0.05, rely=0.64)  # Ajusta la posición
# Fin del apartado de Periodo Incompleto

def Calculos():
    Tabla.bind("<Button-1>", toggle_checkbox)

    columnas = [0] * 6
    conteo = [0] * 6  # Para contar cuántos valores válidos hay en cada columna

    for row in tabla.table:
        for col, entry in enumerate(row):
            try:
                if isinstance(entry, ctk.CTkEntry):  # Si es un Entry, obtener el valor con .get()
                    value = entry.get().strip()
                elif isinstance(entry, ctk.CTkLabel):  # Si es un Label, obtener el valor con .cget("text")
                    value = entry.cget("text").strip()
                else:
                    continue  # Si no es ninguno de los dos, ignorarlo

                if value:  # Verifica que el campo no esté vacío
                    columnas[col] += float(value)  # Convertir el valor a float y sumarlo
                    conteo[col] += 1  # Aumentar conteo de valores válidos
            except ValueError:
                pass  # Ignorar valores no numéricos

    # Calcular promedios con manejo de división por cero
    promedio_voltaje = columnas[2] / conteo[2] if conteo[2] > 0 else 0
    promedio_hrs_uso = columnas[4] / conteo[4] if conteo[4] > 0 else 0

    # Asignar valores a las etiquetas correspondientes
    corriente.configure(text=f"{columnas[1]:.2f}")
    voltaje.configure(text=f"{promedio_voltaje:.1f}")
    potencia.configure(text=f"{columnas[3]:.3f}")
    HrsUso.configure(text=f"{promedio_hrs_uso:.2f}")
    cpd.configure(text=f"{columnas[5]:.4f}")

    # Limpiar la tabla antes de insertar nuevos datos
    Tabla.delete(*Tabla.get_children())

    # Variables para obtener los cálculos
    suma_dias = 0
    suma_kWh_total = 0
    suma_kWh_total_DF = 0
    suma_kWh_total_D = 0

    MedSelect = selected_medidor.get()

    tarifa = BD.tarifa(RPU, MedSelect)
    lbl25.configure(text=tarifa[0])

    # Mostrar los datos en la tabla

    desde_año, desde_mes, desde_mes_palabra, desde_dia, hasta_año, hasta_mes, hasta_mes_palabra, hasta_dia = periodo_selector.get_range()
    desde = f"{desde_año}-{desde_mes}-{desde_dia}"
    hasta = f"{hasta_año}-{hasta_mes}-{hasta_dia}"
    
    for row in BD.mostrardatos(RPU, desde, hasta):
        FECHA, DESDE, HASTA, CONSUMO = row

        # Convertir CONSUMO a entero si es necesario
        if isinstance(CONSUMO, str):
            CONSUMO = int(CONSUMO)

        # Formatear la fecha para mostrar los últimos 2 dígitos del año y los 2 dígitos del mes
        FECHA_FORMATO = FECHA.strftime('%y%m')

        # Formatear las fechas desde y hasta para mostrar solo la fecha y no la hora
        DESDE_FORMATO = DESDE.strftime('%Y-%m-%d')
        HASTA_FORMATO = HASTA.strftime('%Y-%m-%d')

        # Calcular nuevas columnas
        DIAS = (HASTA - DESDE).days

        # Insertar en la tabla en orden cronológico
        Tabla.insert("", "end", values=(FECHA_FORMATO, DESDE_FORMATO, HASTA_FORMATO, DIAS, CONSUMO, " ", " ", "✓"))
    
    totalDias.configure(text=suma_dias)
    totalCFRkwh.configure(text=suma_kWh_total)
    totalDFkwh.configure(text=suma_kWh_total_DF)
    totalDkwh.configure(text=suma_kWh_total_D)
    lbl28.configure(text=suma_kWh_total_D)

    # Diccionario de anomalias
    Anomalias = {
        "EF01": "DIRECTO CON ESTIMACION ERRÓNEA",
        "EF02": "TARIFA INCORRECTA",
        "EF03": "SERVICIO CON LECTURAS ESTIMADAS",
        "EF04": "SERVICIO CON MEDICIÓN Y NO SE FACTURA",
        "EF05": "SERVICIO DE CONCURRENCIA DE TARIFAS",
        "EF06": "LECTURA MAL TOMADA",
        "EF07": "CARGA MAYOR A LA CONTRATADA",
        "EF08": "MULTIPLICADOR DE LECTURAS ERRÓNEO",
        "EF09": "CLAVE DE TIPO DE SUMINISTRO ERRÓNEA",

        "FM01": "SERVICIO DIRECTO",
        "FM02": "MULTIPLICADOR DE LECTURAS ERRÓNEO",
        "FM03": "MEDIDOR MAL CONECTADO",
        "FM04": "BOBINA DE POTENCIAL ABIERTA",
        "FM05": "MEDIDOR QUEMADO",
        "FM06": "TRANSFORMADOR DE INSTRUMENTO DAÑADO",
        "FM07": "TRANSFORMADOR DE INSTRUMENTO MAL CONTECTADO",
        "FM08": "CIRCUTIO DE CORRIENTE DE T.C A TIERRA",
        "FM09": "REGISTRO SIN ENGRANAR",
        "FM10": "NÚMERO DE MEDIDOR DIFERENTE AL DE FACTURACION",
        "FM11": "CLAVE DE TIPO DE SUMINISTRO MAL ERRÓNEA",
        "FM12": "MEDICION INADECUADA AL CIRCUITO",
        "FM13": "MEDIDOR DESTRUIDO",
        "FM14": "ALAMBRADO SECUNDARIO DAÑADO",
        "FM15": "SECUDNARIO DE T.C. EN CORTO CIRCUITO",
        "FM16": "MEDIDOR MAL PROGRAMADO",
        "FM17": "REGISTRO ATORADO",
        "FM18": "DISCO ATORADO",

        "UI01": "SELLOS VIOLADOS",
        "UI02": "SELLO FALSIFICADO",
        "UI03": "SERVICIO DIRECTO SIN CONTRATO",
        "UI04": "CARGA CONECTADA ANTES DE LA MEDICIÓN",
        "UI05": "MANECILLAS INVERTIDAS",
        "UI06": "DISCO ATORADO",
        "UI07": "AJUSTE DE MEDIDORES MOVIDOS",
        "UI08": "RETORNO DE CORRIENTE ABIERTO",
        "UI09": "INTERVENCIÓN DE LAS CONEXIONES DEL EQUIPO",
        "UI10": "ALDABAS DE POTENCILA ABIERTAS",
        "UI11": "MEDIDOR INVERTIDO",
        "UI12": "TERMINALES DE BASE ENCHUFES PUENTEADAS",
        "UI13": "SIN CONTRATO CON MEDICIÓN INSTALADA",
        "UI14": "PASA ENERGÍA A OTRO DOMICILIO",
        "UI15": "ALTERACIÓN A LA PROGRAMACIÓN DEL MEDIDOR",
        "UI16": "RETIRA SU MEDIDOR Y COLOCA OTRO QUE NO FACTURA",
        "UI17": "REGISTRO DEL MEDIDOR ALTERADO O CAMBIADO",
        "UI18": "OTROS USOS INDEBIDOS"
    }

    codigo_anomalia = anomalia.get()
    nombre_anomalia = Anomalias.get(codigo_anomalia, "Desconocida")  # Si no está en el diccionario, muestra "Desconocida"
    lbl33.configure(text=nombre_anomalia)

    ultimo_periodo_incompleto["desde"] = 0
    ultimo_periodo_incompleto["hasta"] = 0
    ultimo_resumen_incompleto["dias"] = 0
    ultimo_resumen_incompleto["df"] = 0
    ultimo_resumen_incompleto["d"] = 0

    agregar.configure(state="disabled")
    btnConfirmar.configure(state="normal")

    return suma_dias, suma_kWh_total, suma_kWh_total_DF, suma_kWh_total_D, rpu, desde, hasta

cargar = ctk.CTkButton(calculos, text="Calcular",fg_color="#2b2b2b", bg_color="Gray", width=100, height=40, font=("Arial", 14, "bold"), hover_color="Green", command=Calculos)
cargar.place(relx=0.6, rely=0.53, relwidth=0.2, relheight=0.06)

def confirmar_calculo():
     # Variables acumuladoras
    suma_dias = 0
    suma_kWh_total = 0
    suma_kWh_total_DF = 0
    suma_kWh_total_D = 0

    Tabla.unbind("<Button-1>")  # Desactiva el click sobre la tabla para evitar más selecciones

    try:
        CPD = float(cpd.cget("text"))
    except ValueError:
        CPD = 0  

    for item in Tabla.get_children():
        valores = list(Tabla.item(item, "values"))

        if valores[7] == "✓":
            try:
                FECHA_FORMATO = valores[0]
                DESDE = datetime.strptime(valores[1], '%Y-%m-%d')
                HASTA = datetime.strptime(valores[2], '%Y-%m-%d')
                DIAS = int(valores[3])
                CONSUMO = int(valores[4])  # Ya viene en la tabla

                CONSUMO_DF = math.ceil(DIAS * CPD)  # Redondear hacia arriba o abajo
                CONSUMO_D = math.ceil(CONSUMO_DF - CONSUMO)  # Redondear hacia arriba o abajo

                # Reemplazar valores, manteniendo los anteriores y agregando los nuevos
                Tabla.item(item, values=(FECHA_FORMATO, DESDE.strftime('%Y-%m-%d'), HASTA.strftime('%Y-%m-%d'), DIAS, CONSUMO, CONSUMO_DF, CONSUMO_D, "✓" ))

                # Acumular en variables
                suma_dias += DIAS
                suma_kWh_total += CONSUMO
                suma_kWh_total_DF += CONSUMO_DF
                suma_kWh_total_D += CONSUMO_D

            except Exception as e:
                print(f"Error al calcular fila: {e}")
        else:
            Tabla.delete(item)  # Eliminar filas no seleccionadas

    # Actualizar etiquetas con resultados
    totalDias.configure(text=suma_dias)
    totalCFRkwh.configure(text=suma_kWh_total)
    totalDFkwh.configure(text=suma_kWh_total_DF)
    totalDkwh.configure(text=suma_kWh_total_D)
    lbl28.configure(text=suma_kWh_total_D)

    # Obtener fechas del periodo_selector
    items = Tabla.get_children()
    if not items:
        messagebox.showerror("Error", "⚠ No hay datos en la tabla para validar.")
        return False
    
    # Tomar el primer y último registro de la tabla
    primer_registro = Tabla.item(items[0], "values")  # Primer registro
    ultimo_registro = Tabla.item(items[-1], "values")  # Último registro

    desde_año_txt = datetime.strptime(primer_registro[1], "%Y-%m-%d").strftime("%Y")
    desde_mes_palabra_txt = datetime.strptime(primer_registro[1], "%Y-%m-%d").strftime("%B")
    desde_dia_txt = datetime.strptime(primer_registro[1], "%Y-%m-%d").strftime("%d")

    hasta_año_txt = datetime.strptime(ultimo_registro[2], "%Y-%m-%d").strftime("%Y")
    hasta_mes_palabra_txt = datetime.strptime(ultimo_registro[2], "%Y-%m-%d").strftime("%B")
    hasta_dia_txt = datetime.strptime(ultimo_registro[2], "%Y-%m-%d").strftime("%d")

    lbl31.configure(text=f"Del {desde_dia_txt} de {desde_mes_palabra_txt} de {desde_año_txt} al {hasta_dia_txt} de {hasta_mes_palabra_txt} de {hasta_año_txt}")

    btnConfirmar.configure(state="disabled")
    agregar.configure(state="normal")

# Guardar último período incompleto agregado
ultimo_periodo_incompleto = {"desde": None, "hasta": None}
ultimo_resumen_incompleto = {"dias": 0, "df": 0, "d": 0}

def Agregar():
    global ultimo_periodo_incompleto, ultimo_resumen_incompleto

    # Obtener fechas del periodo incompleto
    incompleto_desde_año, incompleto_desde_mes, _, incompleto_desde_dia, \
    incompleto_hasta_año, incompleto_hasta_año_formato, \
    incompleto_hasta_mes, _, incompleto_hasta_dia = periodo_selector_incompleto.get_range()

    desde_incompleto = f"{incompleto_desde_año}-{incompleto_desde_mes}-{incompleto_desde_dia}"
    hasta_incompleto = f"{incompleto_hasta_año}-{incompleto_hasta_mes}-{incompleto_hasta_dia}"
    fecha_incompleto = f"{incompleto_hasta_año_formato}{incompleto_hasta_mes}"

    # 🔒 Validar si es el mismo periodo ya agregado
    if (ultimo_periodo_incompleto.get("desde") == desde_incompleto and
        ultimo_periodo_incompleto.get("hasta") == hasta_incompleto):
        messagebox.showinfo("Sin cambios", "⚠ Mismo periodo incompleto agregado, no se ha modificado la tabla.")
        return

    # 🔄 Restar resumen anterior si existía
    try:
        suma_actual_dias = int(totalDias.cget("text")) - int(ultimo_resumen_incompleto.get("dias", 0))
        suma_actual_df = int(totalDFkwh.cget("text")) - int(ultimo_resumen_incompleto.get("df", 0))
        suma_actual_d = int(totalDkwh.cget("text")) - int(ultimo_resumen_incompleto.get("d", 0))
    except Exception as e:
        print("Error en resta de resumen:", e)
        suma_actual_dias = int(totalDias.cget("text"))
        suma_actual_df = int(totalDFkwh.cget("text"))
        suma_actual_d = int(totalDkwh.cget("text"))


    # 🔻 Eliminar fila anterior (si existe)
    for item_id in Tabla.get_children():
        valores = Tabla.item(item_id)["values"]
        if valores and valores[7] == " ":
            Tabla.delete(item_id)
            break

    # ✅ Validar fechas antes de insertar
    if not validar_fecha(Tabla, desde_incompleto, hasta_incompleto):
        return

    # Calcular nuevos valores
    desde_dt = datetime.strptime(desde_incompleto, "%Y-%m-%d").date()
    hasta_dt = datetime.strptime(hasta_incompleto, "%Y-%m-%d").date()
    dias = (hasta_dt - desde_dt).days

    try:
        CPD = float(cpd.cget("text"))
    except ValueError:
        CPD = 0

    consumo_DF = math.ceil(dias * CPD)
    consumo_D = consumo_DF

    # ➕ Insertar nueva fila
    insertar_en_orden(Tabla, fecha_incompleto, (
        fecha_incompleto, desde_incompleto, hasta_incompleto, dias, 0,
        consumo_DF, consumo_D, " "))

    # 🧮 Actualizar sumas
    totalDias.configure(text=suma_actual_dias + dias)
    totalDFkwh.configure(text=suma_actual_df + consumo_DF)
    totalDkwh.configure(text=suma_actual_d + consumo_D)
    lbl28.configure(text=suma_actual_d + consumo_D)

    # 💾 Guardar nuevo estado
    ultimo_periodo_incompleto["desde"] = desde_incompleto
    ultimo_periodo_incompleto["hasta"] = hasta_incompleto

    ultimo_resumen_incompleto["dias"] = dias
    ultimo_resumen_incompleto["df"] = consumo_DF
    ultimo_resumen_incompleto["d"] = consumo_D

    # 🗓 Actualizar periodo visible
    fechas = []
    for item_id in Tabla.get_children():
        valores = Tabla.item(item_id)["values"]
        if len(valores) >= 3:
            fechas.append(datetime.strptime(valores[1], "%Y-%m-%d").date())
            fechas.append(datetime.strptime(valores[2], "%Y-%m-%d").date())

    if fechas:
        inicio = min(fechas)
        fin = max(fechas)
        lbl31.configure(
            text=f"Del {inicio.strftime('%d')} de {inicio.strftime('%B')} de {inicio.year} "
                 f"al {fin.strftime('%d')} de {fin.strftime('%B')} de {fin.year}")

def insertar_en_orden(tabla, nueva_fecha, valores):
    """Inserta una fila en orden cronológico dentro de la tabla"""
    filas = tabla.get_children()
    posicion = None

    # Convertir nueva fecha a número para comparación
    nueva_fecha = int(nueva_fecha)

    for fila in filas:
        fila_valores = tabla.item(fila, "values")
        fecha_existente = int(fila_valores[0])  # Primera columna es la fecha

        if nueva_fecha < fecha_existente:
            posicion = fila  # Insertar antes de esta fila
            break

    if posicion:
        tabla.insert(parent=tabla.parent(posicion), index=tabla.index(posicion), values=valores) # 🔹 Insertar en orden
    else:
        tabla.insert("", "end", values=valores)  # Si no encontró, insertar al final

def validar_fecha(treeview, periodo_desde, periodo_hasta):
    # Asegurar que las fechas sean objetos date
    if isinstance(periodo_desde, str):
        periodo_desde = datetime.strptime(periodo_desde, "%Y-%m-%d").date()
    if isinstance(periodo_hasta, str):
        periodo_hasta = datetime.strptime(periodo_hasta, "%Y-%m-%d").date()

    for item_id in treeview.get_children():
        item = treeview.item(item_id)
        valores = item["values"]

        if len(valores) >= 3:
            try:
                fecha_existente_desde = datetime.strptime(valores[1], "%Y-%m-%d").date()
                fecha_existente_hasta = datetime.strptime(valores[2], "%Y-%m-%d").date()
            except Exception as e:
                print(f"Error procesando fechas de fila: {valores}, error: {e}")
                continue

            # Validaciones de solapamiento
            if (
                fecha_existente_desde < periodo_desde < fecha_existente_hasta or
                fecha_existente_desde < periodo_hasta < fecha_existente_hasta or
                periodo_desde < fecha_existente_desde and periodo_hasta > fecha_existente_hasta
            ):
                messagebox.showerror(
                    "Error",
                    f"⚠ Las fechas seleccionadas ({periodo_desde} a {periodo_hasta}) se solapan con un período existente ({fecha_existente_desde} a {fecha_existente_hasta})."
                )
                return False

    return True  # Si pasó por todos sin solapamiento

def toggle_checkbox(event):
    item_id = Tabla.identify_row(event.y)
    column = Tabla.identify_column(event.x)

    if column == "#8" and item_id:  # "#8" es la columna SELECCION (índice 1-based)
        valores = list(Tabla.item(item_id, "values"))
        valores[7] = "✓" if valores[7] != "✓" else "✗"
        Tabla.item(item_id, values=valores)

agregar = ctk.CTkButton(calculos, text="Agregar",fg_color="#2b2b2b", bg_color="Gray", width=100, height=40, font=("Arial", int(fuente_base * 0.6), "bold"), hover_color="#FFC300", command=Agregar, state="disabled")
agregar.place(relx=0.4, rely=0.79, relwidth=0.2, relheight=0.06)
# Fin de Seccion de Calculos

# Seccion de Tabla de resultados
SeccionResultados = customtkinter.CTkFrame(ventana, fg_color="#C0C0C0", bg_color="#C0C0C0")
SeccionResultados.place(relx=0.35, rely=0.27, relwidth=0.65, relheight=0.75)

ResultadosTabla = customtkinter.CTkFrame(ventana, fg_color="Black", bg_color="Black")
ResultadosTabla.place(relx=0.43, rely=0.3, relwidth=0.4925, relheight=0.075)

def abrir_ventana_reporte():
    # Obtener el tamaño de la pantalla
    screen_width = ventana.winfo_screenwidth()
    screen_height = ventana.winfo_screenheight()

    ventana_reporte = customtkinter.CTkToplevel()
    ventana_reporte.title("Editor de Reporte")
    ventana_reporte.geometry(f"{screen_width}x{screen_height}")
    ventana_reporte.config(bg="#C0C0C0")

    estiloTablaTitulo = { 
        "fg_color": ("#C0C0C0"), 
        "bg_color": ("#C0C0C0"), 
        "text_color": ("White"),
        "font": ("Arial", int(fuente_base * 0.8),"bold")
    }

    estiloCamposTexto = { 
        "fg_color": ("White"), 
        "bg_color": ("White"), 
        "text_color": ("Black"),
        "font": ("Arial", int(fuente_base * 0.6),"bold")
    }

    titulo = customtkinter.CTkLabel(ventana_reporte, text="Vista Previa del Reporte", **estiloTablaTitulo)
    titulo.pack(pady=10)

    campos = {
        "Descripción": None,
        "Método": None,
        "Giro": None,
        "Periodo": None,
        "Persona": None,
        "Situación": None
    }

    columnas = 2
    anchura_textbox = 600
    altura_textbox = 150
    separacion_horizontal = 0.5  # separa columnas
    separacion_vertical = 0.3   # separa filas

    for i, campo in enumerate(campos.keys()):
        fila = i // columnas
        columna = i % columnas

        # Posiciones relativas
        relx = 0.05 + columna * separacion_horizontal
        rely = 0.05 + fila * separacion_vertical

        # Etiqueta
        label = customtkinter.CTkLabel(ventana_reporte, text=campo + ":", anchor="w", **estiloTablaTitulo)
        label.place(relx=relx, rely=rely)

        # Caja de texto
        textbox = customtkinter.CTkTextbox(ventana_reporte, width=anchura_textbox, height=altura_textbox, **estiloCamposTexto)
        textbox.place(relx=relx, rely=rely + 0.05)

        campos[campo] = textbox

        # Switch para incluir la gráfica
        incluir_grafica_var = customtkinter.BooleanVar(value=False)

        switch_grafica = customtkinter.CTkSwitch(
            ventana_reporte,
            text="Incluir gráfica en el PDF",
            variable=incluir_grafica_var,
            onvalue=True,
            offvalue=False,
            switch_width=50,
            switch_height=25,
            font=("Arial", int(fuente_base * 0.6), "bold"),
            text_color="black",
            fg_color = "White",
            progress_color="Green",
            bg_color="#C0C0C0"
        )
        switch_grafica.place(relx=0.55, rely=0.9)  # Ajusta `rely` según tu diseño

    def confirmar_generacion():
        descripcion = campos["Descripción"].get("1.0", "end").strip()
        metodo = campos["Método"].get("1.0", "end").strip()
        giro = campos["Giro"].get("1.0", "end").strip()
        periodo = campos["Periodo"].get("1.0", "end").strip()
        persona = campos["Persona"].get("1.0", "end").strip()
        situacion = campos["Situación"].get("1.0", "end").strip()

        if not descripcion or not metodo or not giro or not persona or not situacion:
            messagebox.showerror("Error", "Por favor completa todos los campos antes de generar el PDF.")
            return

        # Recolectar datos automáticos desde la GUI principal
        nombre = Nomrpu.cget("text")
        medidor = selected_medidor.get()
        tarifa = lbl25.cget("text")
        rpu = RPU
        anomalia_texto = anomalia.get() + " " + lbl33.cget("text")
        periodo_ajuste = periodo + " " + lbl31.cget("text")
        rpe = RPE
        nomrpe = Nomrpe.cget("text")
        tipo_tabla = "CARGA"

        datos_tabla = tabla.obtener_datos_tabla()
        corriente_total = corriente.cget("text")
        voltaje_total = voltaje.cget("text")
        potencia_total = potencia.cget("text")
        hruso_total = HrsUso.cget("text")
        cpd_total = cpd.cget("text")

        cpd_data = [["Fase", "Corriente","Voltaje", "Potencia", "Horas de uso", "C.P.D."]]  # encabezado

        for fila in datos_tabla:
            # Asegúrate de que todos los valores sean strings
            cpd_data.append([str(fila[0]), str(fila[1]), str(fila[2]), str(fila[3]), str(fila[4]), str(fila[5])])

        # Agregar fila de totales
        cpd_data.append([
            "TOTAL",
            str(corriente_total),
            str(voltaje_total),
            str(potencia_total),
            str(hruso_total),
            f"{float(cpd_total):.4f}"
        ])

        # Selección de archivo
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if not filepath:
            return  # Cancelado por el usuario
        
        incluir_grafica = incluir_grafica_var.get()

        fechas_grafica, consumos_grafica, consumos_df_grafica = [], [], []

        if incluir_grafica:
            try:
                for item in Tabla.get_children():
                    fila = Tabla.item(item)["values"]
                    if len(fila) >= 6:
                        fechas_grafica.append(int(fila[0]))
                        consumos_grafica.append(float(fila[4]))
                        consumos_df_grafica.append(float(fila[5]))
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo extraer la información para la gráfica: {e}")
                return

        generar_reporte_pdf(
            nombre, direccion, medidor, tarifa, rpu, anomalia_texto,
            descripcion, metodo, giro, periodo_ajuste, persona, situacion,
            cpd_data, tipo_tabla, rpe, nomrpe, filepath, "", "", incluir_grafica=incluir_grafica,
            fechas=fechas_grafica, consumos=consumos_grafica, consumos_df=consumos_df_grafica
        )

        ventana_reporte.destroy()
        messagebox.showinfo("Reporte generado", "📄 El PDF se ha guardado exitosamente")

    boton_generar = customtkinter.CTkButton(ventana_reporte, text="Generar", command=confirmar_generacion, fg_color="#e74c3c", hover_color="#cb4335", bg_color="#C0C0C0", font=("Arial", int(fuente_base * 0.55), "bold"))
    boton_generar.place(relx=0.45, rely=0.9, relwidth=0.08, relheight=0.04)

def exportar_a_excel():
    try:
        # Abrir plantilla
        plantilla_path = "Plantilla Carga Instantanea.xlsx"
        wb = load_workbook(plantilla_path)
        ws = wb["X F. A."]

        # Datos principales
        nombre = Nomrpu.cget("text")
        agen = agencia.cget("text")
        medidor = selected_medidor.get()
        tarifa = lbl25.cget("text")
        rpu = RPU
        cuent = cuenta.cget("text")
        anomalia = lbl33.cget("text")
        periodo_ajuste = lbl31.cget("text")
        nomrpe = Nomrpe.cget("text")
        
        corriente_total = corriente.cget("text")
        voltaje_total = voltaje.cget("text")
        potencia_total = potencia.cget("text")
        hruso_total = HrsUso.cget("text")
        cpd_total = cpd.cget("text")

        # Insertar en celdas fijas
        ws["F10"] = nombre
        ws["C9"] = rpu
        ws["F9"] = medidor
        ws["K5"] = datetime.now().strftime("%A, %d %B, %Y")
        ws["J10"] = agen
        ws["C10"] = cuent

        fila_encabezado = 16  # Donde se empezaran a colocar los datos

        # -----------------------------------
        # 📋 INSERTAR DATOS DINÁMICOS DE LA TABLA
        # -----------------------------------

        datos_tabla = tabla.obtener_datos_tabla()
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for i, fila in enumerate(datos_tabla):
            for j, valor in enumerate(fila):
                celda = ws.cell(row=fila_encabezado + i, column=5 + j, value=valor)
                celda.font = Font(name="Arial", size=24)
                celda.alignment = Alignment(horizontal="center", vertical="center")
                celda.border = thin_border
        
        # Insertar totales al final de la tabla
        total_cpd_tabla = fila_encabezado + len(datos_tabla)
        fuente_totales = Font(name="Arial", size=24, bold=True)  # Arial, 24pt, negritas
        alineacion = Alignment(horizontal="center", vertical="center")

        ws[f"E{total_cpd_tabla}"] = "TOTAL"
        ws[f"F{total_cpd_tabla}"] = corriente_total
        ws[f"G{total_cpd_tabla}"] = voltaje_total
        ws[f"H{total_cpd_tabla}"] = potencia_total
        ws[f"I{total_cpd_tabla}"] = hruso_total
        ws[f"J{total_cpd_tabla}"] = cpd_total

        for col in range(5, 11):  # Columnas E (5) a J (10)
            cell = ws.cell(row=total_cpd_tabla, column=col)
            cell.border = thin_border
            cell.font = fuente_totales
            cell.alignment = alineacion

        # --------------------------
        # 🔢 TOTAL GENERAL C.P.D.
        # --------------------------

        fila_cpd_total = total_cpd_tabla + 2

        # Estilos comunes
        fondo_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fuente_bold_24 = Font(bold=True, name="Arial", size=24)

        # Aplicar valor, estilo y fondo a las celdas del C.P.D.
        celda_cpd_texto = ws.cell(row=fila_cpd_total, column=7, value="C.P.D. Total:")
        celda_cpd_texto.font = fuente_bold_24
        celda_cpd_texto.fill = fondo_amarillo
        celda_cpd_texto.border = thin_border
        celda_cpd_texto.alignment = Alignment(horizontal="center", vertical="center")

        celda_cpd_valor = ws.cell(row=fila_cpd_total, column=8, value=cpd_total)
        celda_cpd_valor.font = fuente_bold_24
        celda_cpd_valor.fill = fondo_amarillo
        celda_cpd_valor.border = thin_border
        celda_cpd_valor.alignment = Alignment(horizontal="center", vertical="center")

        # --------------------------
        # 🟨 ENCABEZADO DE LA TABLA
        # --------------------------
        
        # Combinar celdas para título principal
        # Definir fila para encabezado
        fila_titulo_principal = fila_cpd_total + 2
        ws.row_dimensions[fila_titulo_principal].height = 62 # Ajustar altura de filas en puntos (62 equivale aprox. a 124 px en Excel)

        # 🟡 Título principal: "CONSUMOS FACTURADOS REFLEJADOS EN SICOM"
        ws.merge_cells(start_row=fila_titulo_principal, start_column=5, end_row=fila_titulo_principal, end_column=9)
        celda_sicom = ws.cell(row=fila_titulo_principal, column=5)
        celda_sicom.value = "CONSUMOS FACTURADOS REFLEJADOS EN SICOM"
        celda_sicom.font = Font(name="Arial", size=24, bold=True)
        celda_sicom.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        celda_sicom.alignment = Alignment(horizontal="center", vertical="center")
        # Aplicar borde a todas las celdas del rango combinado
        for col in range(5, 10):  # Columnas E (5) a I (9)
            celda = ws.cell(row=fila_titulo_principal, column=col)
            celda.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # 🟩 Título "DEBIO FACTURAR"
        ws.merge_cells(start_row=fila_titulo_principal, start_column=10, end_row=fila_titulo_principal, end_column=10)
        celda_debio = ws.cell(row=fila_titulo_principal, column=10)
        celda_debio.value = "DEBIO\nFACTURAR"
        celda_debio.font = Font(name="Arial", size=24, bold=True)
        celda_debio.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        celda_debio.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda_debio.border = thin_border

        # 🟧 Título "DIFERENCIA"
        ws.merge_cells(start_row=fila_titulo_principal, start_column=11, end_row=fila_titulo_principal, end_column=11)
        celda_diferencia = ws.cell(row=fila_titulo_principal, column=11)
        celda_diferencia.value = "DIFERENCIA"
        celda_diferencia.font = Font(name="Arial", size=24, bold=True)
        celda_diferencia.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        celda_diferencia.alignment = Alignment(horizontal="center", vertical="center")
        celda_diferencia.border = thin_border

        # Títulos de columnas
        titulos = ["Fecha", "Desde", "Hasta", "Días", "kWh Total", "kWh Total", "kWh Total"]
        fila_titulos = fila_cpd_total + 3
        ws.row_dimensions[fila_titulos].height = 62 # Ajustar altura de filas en puntos (62 equivale aprox. a 124 px en Excel)

        for col, titulo in enumerate(titulos, start=5):
            celda = ws.cell(row=fila_titulos, column=col, value=titulo)
            celda.font = Font(name="Arial", size=24, bold=True)
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # -----------------------------------
        # 📝 OTROS DATOS AL FINAL DEL DOCUMENTO
        # -----------------------------------

        # Eliminar datos anteriores (opcional)
        fila_inicio = fila_titulos + 1 # Fila inicial para poner los datos de la tabla de datos
        for row in ws.iter_rows(min_row=fila_inicio, max_row=fila_inicio + 100, min_col=5, max_col=11):
            for cell in row:
                cell.value = None

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Copiar valores de la tabla sin copiar estilos
        for i, item_id in enumerate(Tabla.get_children()):
            valores = Tabla.item(item_id)["values"]
            for j, valor in enumerate(valores[:7]):  # Solo los primeros 7 valores
                col = 5 + j  # Comienza en la columna E (5)
                cell = ws.cell(row=fila_inicio + i, column=col)
                cell.value = valor
                cell.border = thin_border  # 👈 Aquí aplicamos el borde
        
        # Insertar totales al final de la tabla
        fila_total = fila_inicio + len(Tabla.get_children())
        relleno_totales = PatternFill(fill_type="solid", fgColor="FFFF00")  # Amarillo
        fuente_totales = Font(name="Arial", size=24, bold=True)  # Arial, 24pt, negritas

        ws[f"G{fila_total}"] = "TOTAL"
        ws[f"H{fila_total}"] = totalDias.cget("text")
        ws[f"I{fila_total}"] = totalCFRkwh.cget("text")
        ws[f"J{fila_total}"] = totalDFkwh.cget("text")
        ws[f"K{fila_total}"] = totalDkwh.cget("text")

        for col in range(5, 12):  # Columnas E (5) a K (11)
            cell = ws.cell(row=fila_total, column=col)
            cell.border = thin_border
            cell.fill = relleno_totales
            cell.font = fuente_totales
        
        # Insertar valores inferiores dinámicamente
        fila_base = fila_inicio + len(Tabla.get_children()) + 2

        ws[f"F{fila_base}"] = "Tarifa del servicio"
        ws[f"F{fila_base}"].alignment = Alignment(horizontal="left")
        ws[f"H{fila_base}"] = tarifa
        ws[f"H{fila_base}"].alignment = Alignment(horizontal="center")
        ws[f"I{fila_base}"] = "Suministro en BAJA-BAJA"
        ws[f"I{fila_base}"].alignment = Alignment(horizontal="left")

        ws[f"F{fila_base + 2}"] = "Recuperación"
        ws[f"F{fila_base}"].alignment = Alignment(horizontal="left")
        ws[f"H{fila_base + 2}"] = totalCFRkwh.cget("text")
        ws[f"H{fila_base}"].alignment = Alignment(horizontal="center")
        ws[f"I{fila_base + 2}"] = "KWH TOTALES"
        ws[f"I{fila_base}"].alignment = Alignment(horizontal="left")

        ws[f"F{fila_base + 4}"] = "Periodo del ajuste"
        ws[f"F{fila_base}"].alignment = Alignment(horizontal="left")
        ws[f"I{fila_base + 4}"] = periodo_ajuste

        ws[f"G{fila_base + 8}"] = f"Anomalía: {anomalia}"

        ws[f"H{fila_base + 15}"] = f"___________________________________________________________________________"
        ws[f"H{fila_base + 16}"] = f"Ing. {nomrpe}"
        ws[f"H{fila_base + 17}"] = f"Oficina de medición Tula"

        # -----------------------------------
        # 💾 GUARDAR ARCHIVO
        # -----------------------------------

        ruta_salida = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar reporte de Carga Instantanea"
        )
        if ruta_salida:
            wb.save(ruta_salida)
            messagebox.showinfo("Exportación Exitosa", "📄 El archivo Excel se ha guardado correctamente.")

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al generar el Excel:\n{e}")

estiloTablaTitulo = { 
        "fg_color": ("#2fd134"), 
        "bg_color": ("#2fd134"), 
        "text_color": ("White"), 
        "width": 137, 
        "height": 30, 
        "font": ("Arial", int(fuente_base * 0.65),"bold")
}

estiloTablaDatos = { 
        "fg_color": ("White"), 
        "bg_color": ("White"), 
        "text_color": ("Black"), 
        "width": 137, 
        "height": 30, 
        "font": ("Arial", int(fuente_base * 0.65))
}

estiloTablaFinal = { 
        "fg_color": ("White"), 
        "bg_color": ("White"), 
        "text_color": ("Black"), 
        "width": 120, 
        "height": 40, 
        "font": ("Arial", int(fuente_base * 0.65), "bold")
}

# Configurar columnas y filas para que se expandan
for i in range(1, 7):  # columnas 1 a 6
    ResultadosTabla.columnconfigure(i, weight=1)
for i in range(1, 3):  # filas 1 y 2
    ResultadosTabla.rowconfigure(i, weight=1)

col1 = customtkinter.CTkLabel(ResultadosTabla, text="Fase", **estiloTablaTitulo)
col1.grid(row=1, column=1, padx=1, pady=1, sticky="nsew")

fase = customtkinter.CTkLabel(ResultadosTabla, text="Total :", **estiloTablaDatos)
fase.grid(row=2, column=1, padx=1, pady=1, sticky="nsew")

col2 = customtkinter.CTkLabel(ResultadosTabla, text="Corriente", **estiloTablaTitulo)
col2.grid(row=1, column=2, padx=1, pady=1, sticky="nsew")

corriente = customtkinter.CTkLabel(ResultadosTabla, text="0.00", **estiloTablaDatos)
corriente.grid(row=2, column=2, padx=1, pady=1, sticky="nsew")

col3 = customtkinter.CTkLabel(ResultadosTabla, text="Voltaje", **estiloTablaTitulo)
col3.grid(row=1, column=3, padx=1, pady=1, sticky="nsew")

voltaje = customtkinter.CTkLabel(ResultadosTabla, text="0.00", **estiloTablaDatos)
voltaje.grid(row=2, column=3, padx=1, pady=1, sticky="nsew")

col4 = customtkinter.CTkLabel(ResultadosTabla, text="Potencia", **estiloTablaTitulo)
col4.grid(row=1, column=4, padx=1, pady=1, sticky="nsew")

potencia = customtkinter.CTkLabel(ResultadosTabla, text="0.00", **estiloTablaDatos)
potencia.grid(row=2, column=4, padx=1, pady=1, sticky="nsew")

col5 = customtkinter.CTkLabel(ResultadosTabla, text="Horas/Uso", **estiloTablaTitulo)
col5.grid(row=1, column=5, padx=1, pady=1, sticky="nsew")

HrsUso = customtkinter.CTkLabel(ResultadosTabla, text="0.00", **estiloTablaDatos)
HrsUso.grid(row=2, column=5, padx=1, pady=1, sticky="nsew")

col6 = customtkinter.CTkLabel(ResultadosTabla, text="C.P.D", **estiloTablaTitulo)
col6.grid(row=1, column=6, padx=1, pady=1, sticky="nsew")

cpd = customtkinter.CTkLabel(ResultadosTabla, text="0.00", **estiloTablaDatos)
cpd.grid(row=2, column=6, padx=1, pady=1, sticky="nsew")

TablaFinal = customtkinter.CTkFrame(ventana, width=850, height=450, fg_color="Black", bg_color="Black")
TablaFinal.place(relx=0.395, rely=0.4, relwidth=0.564, relheight=0.096)

for i in range(1, 8):  # 7 columnas
    TablaFinal.columnconfigure(i, weight=1)
for i in range(1, 3):  # 2 filas
    TablaFinal.rowconfigure(i, weight=1)

lbl15 = customtkinter.CTkLabel(TablaFinal, text="Consumos facturados\n reflejados en SICOM", fg_color="#fff301", bg_color="#fff301", text_color="Black", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl15.grid(row=1, column=1, columnspan=4, padx=1, pady=1, sticky="nsew")

lbl16 = customtkinter.CTkLabel(TablaFinal, text="Debio\n facturar", fg_color="Orange", bg_color="Orange", text_color="Black", font= ("Arial", int(fuente_base * 0.65),"bold"))
lbl16.grid(row=1, column=5, padx=1, pady=1, sticky="nsew")

lbl17 = customtkinter.CTkLabel(TablaFinal, text="Diferencia", width=120, height=40, fg_color="#ff0000", bg_color="#ff0000", text_color="Black", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl17.grid(row=1, column=6, padx=1, pady=1, sticky="nsew")

lbl35 = customtkinter.CTkLabel(TablaFinal, text="Seleccion", width=120, height=40, fg_color="Gray", bg_color="Gray", text_color="Black", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl35.grid(row=1, column=7, padx=1, pady=1, sticky="nsew")

lbl18 = customtkinter.CTkLabel(TablaFinal, text="Fecha", **estiloTablaFinal)
lbl18.grid(row=2, column=1, padx=1, pady=1, sticky="nsew")

lbl19 = customtkinter.CTkLabel(TablaFinal, text="Fecha \n Desde - Hasta", width=223, height=40, fg_color="White", bg_color="White", text_color="Black", font=("Arial", int(fuente_base * 0.65),"bold"))
lbl19.grid(row=2, column=2, padx=1, pady=1, sticky="nsew")

lbl20 = customtkinter.CTkLabel(TablaFinal, text="Dias", **estiloTablaFinal)
lbl20.grid(row=2, column=3, padx=1, pady=1, sticky="nsew")

lbl21 = customtkinter.CTkLabel(TablaFinal, text="kWh Total", **estiloTablaFinal)
lbl21.grid(row=2, column=4, padx=1, pady=1, sticky="nsew")

lbl22 = customtkinter.CTkLabel(TablaFinal, text="kWh Total", **estiloTablaFinal)
lbl22.grid(row=2, column=5, padx=1, pady=1, sticky="nsew")

lbl23 = customtkinter.CTkLabel(TablaFinal, text="kWh Total",  **estiloTablaFinal)
lbl23.grid(row=2, column=6, padx=1, pady=1, sticky="nsew")

lbl36 = customtkinter.CTkLabel(TablaFinal, text="✓",  **estiloTablaFinal)
lbl36.grid(row=2, column=7, padx=1, pady=1, sticky="nsew")

DatosBD = customtkinter.CTkFrame(ventana,  width=983, height=265, fg_color="Black", bg_color="Black")
DatosBD.pack_propagate(False)
DatosBD.place(relx=0.395, rely=0.5, relwidth=0.573, relheight=0.3)

# Apartado para los datos de la BD

style = ttk.Style()
style.configure("Resultados.Treeview", font=("Arial", int(fuente_base * 0.8)), rowheight=40)

Tabla = ttk.Treeview(DatosBD, style="Resultados.Treeview", columns=("FECHA", "DESDE", "HASTA", "DIAS", "CONSUMO", "CONSUMNO_DF", "CONSUMO_D", "SELECCION"), show="tree")
Tabla.column("#0", width=0, stretch=tk.NO)
Tabla.column("FECHA", anchor=CENTER, width=25)
Tabla.column("DESDE", anchor=CENTER, width=8)
Tabla.column("HASTA", anchor=CENTER, width=8)
Tabla.column("DIAS", anchor=CENTER, width=25)
Tabla.column("CONSUMO", anchor=CENTER, width=25)
Tabla.column("CONSUMNO_DF", anchor=CENTER, width=25)
Tabla.column("CONSUMO_D", anchor=CENTER, width=25)
Tabla.column("SELECCION", anchor=CENTER, width=25)

Tabla.bind("<Button-1>", toggle_checkbox)

for col in ("FECHA", "DESDE", "HASTA", "DIAS", "CONSUMO", "CONSUMNO_DF", "CONSUMO_D", "SELECCION"):
    Tabla.column(col, anchor=tk.CENTER, width=115)
    Tabla.heading(col, text="") 

# Crear Scrollbar
scrollbar = ttk.Scrollbar(DatosBD, orient="vertical", command=Tabla.yview)
Tabla.configure(yscrollcommand=scrollbar.set)

# Posicionar elementos
Tabla.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

TotalBD = customtkinter.CTkFrame(ventana, fg_color="Black", bg_color="Black")
TotalBD.place(relx=0.395, rely=0.8, relwidth=0.563, relheight=0.048)

for i in range(1, 8):  # 7 columnas
    TotalBD.columnconfigure(i, weight=1)
for i in range(1, 3):  # 2 filas
    TotalBD.rowconfigure(i, weight=1)

lblTotal = customtkinter.CTkLabel(TotalBD, text="Total :",  width=344, height=40, fg_color="White", bg_color="White", text_color="Black", font=("Arial", int(fuente_base * 0.65),"bold"))
lblTotal.grid(row=1, column=1, padx=1, pady=1, sticky="nsew")

totalDias = customtkinter.CTkLabel(TotalBD, text="00.0",  **estiloTablaFinal)
totalDias.grid(row=1, column=3, padx=1, pady=1, sticky="nsew")

totalCFRkwh = customtkinter.CTkLabel(TotalBD, text="00.0",  **estiloTablaFinal)
totalCFRkwh.grid(row=1, column=4, padx=1, pady=1, sticky="nsew")

totalDFkwh = customtkinter.CTkLabel(TotalBD, text="00.0",  **estiloTablaFinal)
totalDFkwh.grid(row=1, column=5, padx=1, pady=1, sticky="nsew")

totalDkwh = customtkinter.CTkLabel(TotalBD, text="00.0",  **estiloTablaFinal)
totalDkwh.grid(row=1, column=6, padx=1, pady=1, sticky="nsew")

btnConfirmar = ctk.CTkButton(TotalBD, text="Confirmar",fg_color="White", bg_color="Gray", width=120, height=40, text_color="Black", font=("Arial", int(fuente_base * 0.65), "bold"), hover_color="Gray", command=confirmar_calculo, state="disabled")
btnConfirmar.grid(row=1, column=7, padx=1, pady=1, sticky="nsew")

DatosFinales = { 
        "fg_color": ("#C0C0C0"), 
        "bg_color": ("#C0C0C0"), 
        "text_color": ("Black"), 
        "width": 80, 
        "height": 40, 
        "font": ("Arial", int(fuente_base * 0.65), "bold")
}

lbl24 = customtkinter.CTkLabel(ventana, text="Tarifa del servicio :", **DatosFinales)
lbl24.place(relx=0.5, rely=0.85)

lbl25 = customtkinter.CTkLabel(ventana, text="00", **DatosFinales)
lbl25.place(relx=0.585, rely=0.85)

lbl26 = customtkinter.CTkLabel(ventana, text="Suministro en Baja - Baja", **DatosFinales)
lbl26.place(relx=0.63, rely=0.85)

lbl27 = customtkinter.CTkLabel(ventana, text="Recuperacion :", **DatosFinales)
lbl27.place(relx=0.5, rely=0.88)

lbl28 = customtkinter.CTkLabel(ventana, text="0000", **DatosFinales)
lbl28.place(relx=0.58, rely=0.88)

lbl29 = customtkinter.CTkLabel(ventana, text="kWh Totales", **DatosFinales)
lbl29.place(relx=0.63, rely=0.88)

lbl30 = customtkinter.CTkLabel(ventana, text="Periodo del ajuste :", **DatosFinales)
lbl30.place(relx=0.5, rely=0.91)

lbl31 = customtkinter.CTkLabel(ventana, text="", **DatosFinales)
lbl31.place(relx=0.59, rely=0.91)

lbl32 = customtkinter.CTkLabel(ventana, text="Anomalia :", **DatosFinales)
lbl32.place(relx=0.5, rely=0.94)

lbl33 = customtkinter.CTkLabel(ventana, text="-------------------", **DatosFinales)
lbl33.place(relx=0.59, rely=0.94)

boton_excel = customtkinter.CTkButton(ventana, text="Exportar a Excel", command=exportar_a_excel, fg_color="#4CAF50", hover_color="#45a049", bg_color="#C0C0C0", font=("Arial", int(fuente_base * 0.55), "bold"))
boton_excel.place(relx=0.86, rely=0.88, relwidth=0.08, relheight=0.04)  # Ajusta posición según diseño

btn_pdf = customtkinter.CTkButton(ventana, text="Generar PDF", command=abrir_ventana_reporte, fg_color="#e74c3c", hover_color="#cb4335", bg_color="#C0C0C0", font=("Arial", int(fuente_base * 0.55), "bold"))
btn_pdf.place(relx=0.86, rely=0.94)  # Puedes ajustar relx para moverlo más a la derecha o izquierda
  
ventana.mainloop()